# -*- coding: utf-8 -*-
"""
Created on Thu Mar 28 18:40:37 2019

@author: Tristan O'Hanlon - University of Auckland & Jonathan Rogers

These functions will create a standard model dataset when parsed:
    
reduce_dataset( directory, filename, save_location, start, end, cl_is_fractional=False )

example:
reduce_dataset( 'gfdl_am4', '_Amon_GFDL-AM4_amip_r1i1p1f1_gr1_198001-201412.nc', 'E:/University/University/MSc/Models/climate-analysis/GFDL-AM4', datetime.datetime( 2000, 1, 1 ), datetime.datetime( 2006, 1, 1 ) )

if cl_is_fractional=True - the values get dived by 100
"""
import numpy as np
import os
from netCDF4 import Dataset
from netCDF4 import date2index
import h5py
import math
from scipy import interpolate
from scipy.interpolate import RectBivariateSpline
from scipy import stats
import datetime
from pprint import pprint
from sklearn.impute import SimpleImputer
import constants
from scipy import ndimage as nd
import matplotlib.pyplot as plt
import openpyxl


def reduce_datatable( directory, filename, save_location, start, end, rownum ):

    book = openpyxl.load_workbook('E:/University/University/MSc/Models/climate-analysis/reduced_data/cloud_data.xlsx')
    sheet = book.active
    # def reduce_dataset( directory, filename, save_location, start, end):
    os.chdir( directory )
  
    ################################################---get regional data (time, lat, lon)---################################################

    if 'RCP45-GFDL' in directory or 'RCP45-IPSL' in directory or 'RCP45-MRI' in directory or 'future4K-MIROC6' in directory or 'future4K-MRI' in directory or 'SSP245-MIROC6' in directory:
        with Dataset( 'clt' + filename, 'r') as f:
            raw_lat = constants.extract_data( 'lat', f )
            raw_lon = constants.extract_data( 'lon', f )
            clt = np.mean( constants.extract_data('clt', f ), axis = 0 ) # average over time

        # get indexes for southern ocean -70 to -50 lat

        start_idx = np.abs(raw_lat - (-69.5)).argmin()
        end_idx = np.abs(raw_lat - (-50)).argmin()


        #---get radiative flux data---#

        with Dataset( 'rsdt' + filename, 'r') as f:
            rsdt = np.mean( constants.extract_data('rsdt', f ), axis = 0 ) # average over time
            
        with Dataset( 'rsut' + filename, 'r') as f:
            rsut = np.mean( constants.extract_data('rsut', f ), axis = 0 ) # average over time           

        with Dataset( 'rsutcs' + filename, 'r') as f:
            rsutcs = np.mean( constants.extract_data('rsutcs', f ), axis = 0 ) # average over time           
        
        with Dataset( 'rlut' + filename, 'r') as f:
            rlut = np.mean( constants.extract_data('rlut', f ), axis = 0 ) # average over time           

        with Dataset( 'rlutcs' + filename, 'r') as f:
            rlutcs = np.mean( constants.extract_data('rlutcs', f ), axis = 0 ) # average over time           
        
        if 'MRI' in directory:
            rtmt = ( rsdt ) - ( rsut + rlut ) # https://www4.uwsp.edu/geo/faculty/lemke/geog101/lectures/02_radiation_energy_balance.html
        else:
            with Dataset( 'rtmt' + filename, 'r') as f:
                rtmt = np.mean( constants.extract_data('rtmt', f ), axis = 0 ) # average over time
        
                
        #---create albedo data---#
        
        albedo_reg = rsut / rsdt
        cre_sw_reg = rsutcs - rsut
        cre_lw_reg = rlutcs - rlut
        
        with Dataset( 'cl' + filename, 'r') as f:
            cl = constants.extract_data( 'cl',f )
            cl = np.mean( cl, axis = 0 ) # average over time
            if directory == 'CMIP5-AMIP-CESM1-CAM5':
                pass
            else:
                cl = cl / 100
            if 'GFDL-AM4' in directory or 'IPSL' in directory or 'GFDL-CM4' in directory:
                a = constants.extract_data('ap', f)
                b = constants.extract_data( 'b', f)
            else:
                a = constants.extract_data('a', f )
                b = constants.extract_data('b', f)

            if 'GFDL-AM4' in directory or 'IPSL' in directory or 'GFDL-CM4' in directory:
                pass
            else:
                p0 = np.array(f.variables['p0'][:])
                a = a*p0

        with Dataset( 'clw' + filename, 'r') as f:
            clw = np.nanmean( constants.extract_data( 'clw', f ), axis = 0 ) # average over time

        with Dataset( 'cli' + filename, 'r') as f:
            cli = np.nanmean( constants.extract_data( 'cli', f ), axis = 0 ) # average over time

        with Dataset( 'ps' + filename, 'r') as f:
            ps_lat_lon = np.mean( constants.extract_data( 'ps', f ), axis = 0 ) # average over time

        with Dataset( 'ta' + filename, 'r' ) as f:
            ta = constants.extract_data( 'ta', f )
            ta[ta > 500] = np.nan
            ta = np.nanmean(ta, axis = 0) # average over time  
            plev_t = constants.extract_data('plev', f) / 100
        full_ta_alt_lat = np.nanmean(ta, axis = -1)




    else:
        with Dataset( 'clt' + filename, 'r') as f:
            raw_lat = constants.extract_data( 'lat', f )
            raw_lon = constants.extract_data( 'lon', f )
            clt = np.mean( constants.extract_data_over_time('clt', f, start, end ), axis = 0 ) # average over time

        # get indexes for southern ocean -70 to -50 lat

        start_idx = np.abs(raw_lat - (-69.5)).argmin()
        end_idx = np.abs(raw_lat - (-50)).argmin()


        #---get radiative flux data---#

        with Dataset( 'rsdt' + filename, 'r') as f:
            rsdt = np.mean( constants.extract_data_over_time('rsdt', f, start, end ), axis = 0 ) # average over time
            
        with Dataset( 'rsut' + filename, 'r') as f:
            rsut = np.mean( constants.extract_data_over_time('rsut', f, start, end ), axis = 0 ) # average over time           

        with Dataset( 'rsutcs' + filename, 'r') as f:
            rsutcs = np.mean( constants.extract_data_over_time('rsutcs', f, start, end ), axis = 0 ) # average over time           
        
        with Dataset( 'rlut' + filename, 'r') as f:
            rlut = np.mean( constants.extract_data_over_time('rlut', f, start, end ), axis = 0 ) # average over time           

        with Dataset( 'rlutcs' + filename, 'r') as f:
            rlutcs = np.mean( constants.extract_data_over_time('rlutcs', f, start, end ), axis = 0 ) # average over time           
        
        if 'MRI' in directory:
            rtmt = ( rsdt ) - ( rsut + rlut ) # https://www4.uwsp.edu/geo/faculty/lemke/geog101/lectures/02_radiation_energy_balance.html
        else:
            with Dataset( 'rtmt' + filename, 'r') as f:
                rtmt = np.mean( constants.extract_data_over_time('rtmt', f, start, end ), axis = 0 ) # average over time
        
                
        #---create albedo data---#
        
        albedo_reg = rsut / rsdt
        cre_sw_reg = rsutcs - rsut
        cre_lw_reg = rlutcs - rlut
        
        with Dataset( 'cl' + filename, 'r') as f:
            cl = constants.extract_data_over_time( 'cl',f, start, end )
            cl = np.mean( cl, axis = 0 ) # average over time
            if directory == 'CMIP5-AMIP-CESM1-CAM5':
                pass
            else:
                cl = cl / 100
            if  'IPSL' in directory or 'GFDL-CM4' in directory:
                a = constants.extract_data('ap', f)
                b = constants.extract_data( 'b', f)
            else:
                a = constants.extract_data('a', f )
                b = constants.extract_data('b', f)

            if 'IPSL' in directory or 'GFDL-CM4' in directory:
                pass
            else:
                p0 = np.array(f.variables['p0'][:])
                a = a*p0

        with Dataset( 'clw' + filename, 'r') as f:
            clw = np.nanmean( constants.extract_data_over_time( 'clw', f, start, end ), axis = 0 ) # average over time

        with Dataset( 'cli' + filename, 'r') as f:
            cli = np.nanmean( constants.extract_data_over_time( 'cli', f, start, end ), axis = 0 ) # average over time

        with Dataset( 'ps' + filename, 'r') as f:
            ps_lat_lon = np.mean( constants.extract_data_over_time( 'ps', f, start, end ), axis = 0 ) # average over time

        with Dataset( 'ta' + filename, 'r' ) as f:
            ta = constants.extract_data_over_time( 'ta', f, start, end )
            ta[ta > 500] = np.nan
            ta = np.nanmean(ta, axis = 0) # average over time  
            plev_t = constants.extract_data('plev', f) / 100
        full_ta_alt_lat = np.nanmean(ta, axis = -1)







    a = np.swapaxes(np.swapaxes(np.tile(a, (np.shape(ps_lat_lon)[0], np.shape(ps_lat_lon)[1], 1) ), 0, 2), 1, 2)
    b = np.swapaxes(np.swapaxes(np.tile(b, (np.shape(ps_lat_lon)[0], np.shape(ps_lat_lon)[1], 1) ), 0, 2), 1, 2)
    ps = np.tile(ps_lat_lon, (np.shape(b)[0], 1, 1) )
    p_3D = (a + b*ps) / 100 # in hPa
    
    if directory == 'CMIP6-CESM2-CAM6':
        p_3D = np.flip(p_3D, axis = 0)

    p_alt_lat = np.mean ( p_3D, axis = -1 )
    p = constants.global3DMean(p_3D, raw_lat)
 
    if directory == 'CMIP6-AMIP-CESM2-CAM6':
        cl_alt_lat = np.flip( np.nanmean( cl , axis = -1 ), axis = 0 ) # average over longitude
        clw_alt_lat = np.flip( np.nanmean( clw , axis = -1 ), axis = 0 ) # average over longitude
        cli_alt_lat = np.flip( np.nanmean( cli , axis = -1 ), axis = 0 ) # average over longitude    

    else:
        cl_alt_lat = np.nanmean(cl , axis = -1) # average over longitude
        clw_alt_lat = np.nanmean(clw , axis = -1) # average over longitude
        cli_alt_lat = np.nanmean(cli , axis = -1) # average over longitude      


       #---Approximate missing nan values in temperature data---#  
       
    imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    imp.fit(np.transpose(full_ta_alt_lat))  
    a = imp.transform(np.transpose(full_ta_alt_lat))
    ta_alt_lat = np.transpose(a)    

    if 'GFDL-CM4' in directory :
        interpolated = interpolate.interp2d(plev_t[:-1], raw_lat, np.transpose( ta_alt_lat ), kind = 'cubic')
        ta_alt_lat = np.transpose( interpolated(p, raw_lat) )
    else:
        interpolated = interpolate.interp2d(plev_t, raw_lat, np.transpose( ta_alt_lat ), kind = 'cubic')
        ta_alt_lat = np.transpose( interpolated(p, raw_lat) )

    #---create fractions---#

    clw_frac = ( clw_alt_lat / ( clw_alt_lat + cli_alt_lat ) ) * cl_alt_lat
    cli_frac = ( cli_alt_lat / ( clw_alt_lat + cli_alt_lat ) ) * cl_alt_lat

    imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    imp.fit(np.transpose(clw_frac))  
    a = imp.transform(np.transpose(clw_frac))
    clw_frac = np.transpose(a)   

    imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    imp.fit(np.transpose(cli_frac))  
    a = imp.transform(np.transpose(cli_frac))
    cli_frac = np.transpose(a)   

    # calculate air density at each pressure layer
    air_density_alt_lat = ((p_alt_lat * 100) / (286.9 * ta_alt_lat))

    if 'IPSL-CM6A-LR' in directory:
        air_density_alt_lat = air_density_alt_lat[:-1]

    clwc_alt_lat = ( clw_alt_lat * air_density_alt_lat ) * 1000 # in g/m3
    clic_alt_lat = ( cli_alt_lat * air_density_alt_lat ) * 1000 # in g/m3










    # store model name
    sheet.cell(row=rownum, column=1).value = directory
    # store clt
    sheet.cell(row=rownum, column=2).value = constants.global2DMean(clt, raw_lat)
    # store clt SO
    sheet.cell(row=rownum, column=3).value = constants.global2DMean(clt[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store clw
    sheet.cell(row=rownum, column=4).value = constants.globalalt_latMeanVal(clw_alt_lat, raw_lat)
    # store clw SO
    sheet.cell(row=rownum, column=5).value = constants.globalalt_latMeanVal(clw_alt_lat[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store clwc
    sheet.cell(row=rownum, column=6).value = constants.globalalt_latMeanVal(clwc_alt_lat, raw_lat)
    # store clwc SO
    sheet.cell(row=rownum, column=7).value = constants.globalalt_latMeanVal(clwc_alt_lat[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store clw_frac
    sheet.cell(row=rownum, column=8).value = constants.globalalt_latMeanVal(clw_frac, raw_lat)
    # store clw_frac SO
    sheet.cell(row=rownum, column=9).value = constants.globalalt_latMeanVal(clw_frac[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store cli
    sheet.cell(row=rownum, column=10).value = constants.globalalt_latMeanVal(cli_alt_lat, raw_lat)
    # store cli SO
    sheet.cell(row=rownum, column=11).value = constants.globalalt_latMeanVal(cli_alt_lat[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store clic
    sheet.cell(row=rownum, column=12).value = constants.globalalt_latMeanVal(clic_alt_lat, raw_lat)
    # store clic SO
    sheet.cell(row=rownum, column=13).value = constants.globalalt_latMeanVal(clic_alt_lat[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store cli_frac
    sheet.cell(row=rownum, column=14).value = constants.globalalt_latMeanVal(cli_frac, raw_lat)
    # store cli_frac SO
    sheet.cell(row=rownum, column=15).value = constants.globalalt_latMeanVal(cli_frac[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store albedo
    sheet.cell(row=rownum, column=16).value = constants.global2DMean(albedo_reg, raw_lat)
    # store SO albedo
    sheet.cell(row=rownum, column=17).value = constants.global2DMean(albedo_reg[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store net flux
    sheet.cell(row=rownum, column=18).value = constants.global2DMean(rtmt, raw_lat)
    # store net SO flux
    sheet.cell(row=rownum, column=19).value = constants.global2DMean(rtmt[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store CRE SW
    sheet.cell(row=rownum, column=20).value = constants.global2DMean(cre_sw_reg, raw_lat)
    # store SO CRE SW
    sheet.cell(row=rownum, column=21).value = constants.global2DMean(cre_sw_reg[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store CRE LW
    sheet.cell(row=rownum, column=22).value = constants.global2DMean(cre_lw_reg, raw_lat)
    # store SO CRE LW
    sheet.cell(row=rownum, column=23).value = constants.global2DMean(cre_lw_reg[start_idx:end_idx], raw_lat[start_idx:end_idx])


    book.save('E:/University/University/MSc/Models/climate-analysis/reduced_data/cloud_data.xlsx')

