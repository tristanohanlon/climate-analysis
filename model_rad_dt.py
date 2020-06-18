

import numpy as np
import os
from netCDF4 import Dataset
from netCDF4 import date2index
import h5py
import math
from scipy import interpolate
from scipy.interpolate import RectBivariateSpline
from scipy import stats
import datetime
from pprint import pprint
from sklearn.impute import SimpleImputer
import constants
from scipy import ndimage as nd
import matplotlib.pyplot as plt
import climlab
from climlab.radiation import RRTMG
from matplotlib import rc
rc('text', usetex=True)
from matplotlib.font_manager import FontProperties
fontP = FontProperties()
fontP.set_size('small')
import openpyxl
import xarray as xr



def ensemble_mean( ref_dict ):
    running_sum = np.zeros_like(next(iter(ref_dict.values())))
    for items in ref_dict.values():
        running_sum += items
    mean = running_sum / len(ref_dict)
    return mean


def radiation(start, end, start_dt, end_dt, location, models, label, lat_bnd_1, lat_bnd_2, save_outputs, liquid_r, ice_r ):



    #  Set aerosol and level constants
    o2=0.21
    ccl4=0.
    cfc22=4.8743326488706363e-11
    co2 = 400.e-6

    state = climlab.column_state(num_lev=50)
    lev = state.Tatm.domain.axes['lev'].points*100
    alt = np.flipud(constants.p_to_alt( np.flipud(lev) )  * 1000) # in m
    f = 2 # counter for excel rows
    if label == '2':
        f = 10

    model_CLglobals = {}
    model_CLWPglobals = {}
    model_CIWPglobals = {}
    model_Tglobals = {}
    model_TSglobals = {}
    model_SHglobals = {}
    model_albedos = {}

    model_O3globals = {}
    model_ch4globals = {}
    model_n2oglobals = {}
    model_cfc11globals = {}
    model_cfc12globals = {}
    model_cfc113globals = {}
    
    SWcre_CLdeltas = {}
    LWcre_CLdeltas = {}

    SWcre_LWPdeltas = {}
    LWcre_LWPdeltas = {}

    SWcre_IWPdeltas = {}
    LWcre_IWPdeltas = {}

    SWcre_Tdeltas = {}
    LWcre_Tdeltas = {}

    SWcre_SHdeltas = {}
    LWcre_SHdeltas = {}


    SWcs_CLdeltas = {}
    LWcs_CLdeltas = {}

    SWcs_LWPdeltas = {}
    LWcs_LWPdeltas = {}

    SWcs_IWPdeltas = {}
    LWcs_IWPdeltas = {}

    SWcs_Tdeltas = {}
    LWcs_Tdeltas = {}

    SWcs_SHdeltas = {}
    LWcs_SHdeltas = {}


    SWcre_lr15deltas = {}
    LWcre_lr15deltas = {}

    SWcre_lr45deltas = {}
    LWcre_lr45deltas = {}

    SWcre_lr60deltas = {}
    LWcre_lr60deltas = {}

    SWcre_ir30deltas = {}
    LWcre_ir30deltas = {}

    SWcre_ir80deltas = {}
    LWcre_ir80deltas = {}

    SWcre_ir130deltas = {}
    LWcre_ir130deltas = {}



    colors = ["crimson", "purple", "limegreen", "gold", "blue", "magenta"]

    #  Get global mean valriables for all models and interpolate to common levels
    for name, model in models.items():
        os.chdir( location + 'Data/' + model )

        if 'IPSL' in name or 'MRI' in name or 'MIROC' in name:
            start = start_dt
            end = end_dt

    #----------------------------- Cloud variables -----------------------------#

        #  Take global, annual average

        with xr.open_dataset(constants.variable_to_filename( 'cl' ), decode_times=True) as cl_full:
            cl = cl_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
        weight = np.cos(np.deg2rad(cl.lat)) / np.cos(np.deg2rad(cl.lat)).mean(dim='lat')
        CLglobal = (cl.cl * weight).mean(dim=('lat','lon')) / 100

        if 'CM4' in model or 'IPSL' in model:
            a = cl.ap
            b = cl.b
        else:
            ap = cl.a
            b = cl.b
            p0 = cl.p0
            a = ap*p0

        with xr.open_dataset(constants.variable_to_filename( 'ps' ), decode_times=True) as ps_full:
            ps = ps_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            ps = ps.ps

        p = a + b*ps
        plev = (p * weight).mean(dim=('lat','lon','time'))
        if 'IPSL' in model:
            plev = plev[:-1]


        with xr.open_dataset(constants.variable_to_filename( 'clw' ), decode_times=True) as clw_full:
            clw = clw_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            CLWglobal = (clw.clw * weight).mean(dim=('lat','lon'))

        with xr.open_dataset(constants.variable_to_filename( 'cli' ), decode_times=True) as cli_full:
            cli = cli_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            CLIglobal = (cli.cli * weight).mean(dim=('lat','lon'))

        with xr.open_dataset(constants.variable_to_filename( 'ta' ), decode_times=True) as ta_full:
            ta = ta_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            Tglobal = (ta.ta * weight).mean(dim=('lat','lon'))

        with xr.open_dataset(constants.variable_to_filename( 'hus' ), decode_times=True) as hus_full:
            hus = hus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            SHglobal = (hus.hus * weight).mean(dim=('lat','lon'))  # kg/kg

        with xr.open_dataset(constants.variable_to_filename( 'ts' ), decode_times=True) as ts_full:
            ts = ts_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            TSglobal = (ts.ts * weight).mean(dim=('lat','lon')).values  # kg/kg


                #----------------------------- Aerosols ---------------------------#

        #  Get aerosol values from model files or set default values if they are
        #  not present.
        #  Convert mol/mol to ppm by dividing by 1E-6
        #  Convery CO2 mass to ppm by dividing by atmosphere mass 4.99E18
    
        #  Mole Fraction of Ozone (time, lev, lat, lon)

        check_file = constants.variable_to_filename( 'o3' )
        if check_file == None:
            # use current global mean
            O3global = ensemble_mean(model_O3globals)

        else:
            with xr.open_dataset(constants.variable_to_filename( 'o3'), decode_times=True) as o3_full:
                o3 = o3_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
                O3global = (o3.o3 * weight).mean(dim=('lat','lon'))  # kg/kg


        #  Global Mean Mole Fraction of CH4 (time)

        check_file = constants.variable_to_filename( 'ch4global' )
        if check_file == None:
            ch4 = 1.557889297535934e-6
        else:
            with xr.open_dataset(constants.variable_to_filename( 'ch4global' ), decode_times=True) as ch4_full:
                ch4 = ch4_full.sel(time=slice(start,end))
                ch4 = (ch4.ch4global).mean(dim=('time'))
                ch4 = ch4.values / 1e9


        #  Global Mean Mole Fraction of N2O (time)

        check_file = constants.variable_to_filename( 'n2oglobal' )
        if check_file == None:
            n2o = 3.007494866529774e-7
        else:
            with xr.open_dataset(constants.variable_to_filename( 'n2oglobal' ), decode_times=True) as n2o_full:
                n2o = n2o_full.sel(time=slice(start,end))
                n2o = (n2o.n2oglobal).mean(dim=('time'))
                n2o = n2o.values / 1e9


        #  Global Mean Mole Fraction of CFC11 (time)

        check_file = constants.variable_to_filename( 'cfc11global' )
        if check_file == None:
            cfc11 = 1.680e-10
        else:
            with xr.open_dataset(constants.variable_to_filename( 'cfc11global' ), decode_times=True) as cfc11_full:
                cfc11 = cfc11_full.sel(time=slice(start,end))
                cfc11 = (cfc11.cfc11global).mean(dim=('time'))
                cfc11 = cfc11.values / 1e12


        #  Global Mean Mole Fraction of CFC12 (time)

        check_file = constants.variable_to_filename( 'cfc12global' )
        if check_file == None:
            cfc12 = 2.850e-10
        else:
            with xr.open_dataset(constants.variable_to_filename( 'cfc12global' ), decode_times=True) as cfc12_full:
                cfc12 = cfc12_full.sel(time=slice(start,end))
                cfc12 = (cfc12.cfc12global).mean(dim=('time'))
                cfc12 = cfc12.values / 1e12


        #  Global Mean Mole Fraction of CFC113 (time)

        check_file = constants.variable_to_filename( 'cfc113global' )
        if check_file == None:
            cfc113 = 1.737268993839836e-11
        else:
            with xr.open_dataset(constants.variable_to_filename( 'cfc113global' ), decode_times=True) as cfc113_full:
                cfc113 = cfc113_full.sel(time=slice(start,end))
                cfc113 = (cfc113.cfc113global).mean(dim=('time'))
                cfc113 = cfc113.values / 1e12




        #  interpolate to model pressure levels
        #  Need to 'flipud' because the interpolation routine 
        #  needs the pressure data to be in increasing order
        #  Create a state dictionary with corresponsing number of cl data levels

        if 'CAM6' in model:
            T_plev = np.flipud(Tglobal.plev[:18])
            T_time = np.arange(np.size(Tglobal.time))
            T_ready = np.flip(Tglobal.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(T_plev, T_time, T_ready)
            Tinterp = interpolated(lev, T_time)

            SH_plev = np.flipud(SHglobal.plev[:18])
            SH_ready = np.flip(SHglobal.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(SH_plev, T_time, SH_ready)
            SHinterp = interpolated(lev, T_time)

            O3_plev = np.flipud(O3global.plev[:18])
            O3_ready = np.flip(O3global.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(O3_plev, T_time, O3_ready)
            O3interp = interpolated(lev, T_time)

            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLglobal.values)
            CLinterp = interpolated(lev, T_time)

            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLWglobal.values)
            CLWinterp = interpolated(lev, T_time)

            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLIglobal.values)
            CLIinterp = interpolated(lev, T_time)

        elif 'MRI' in model or 'MIROC' in model:
            O3interp = O3global

            T_plev = np.flipud(Tglobal.plev[:18])
            T_time = np.arange(np.size(Tglobal.time))
            T_ready = np.flip(Tglobal.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(T_plev, T_time, T_ready)
            Tinterp = interpolated(lev, T_time)

            SH_plev = np.flipud(SHglobal.plev[:18])
            SH_ready = np.flip(SHglobal.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(SH_plev, T_time, SH_ready)
            SHinterp = interpolated(lev, T_time)

            CL_ready = np.flip(CLglobal.values, axis = 1)
            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CL_ready)
            CLinterp = interpolated(lev, T_time)

            CLW_ready = np.flip(CLWglobal.values, axis = 1)
            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLW_ready)
            CLWinterp = interpolated(lev, T_time)

            CLI_ready = np.flip(CLIglobal.values, axis = 1)
            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLI_ready)
            CLIinterp = interpolated(lev, T_time)

        else:
            T_plev = np.flipud(Tglobal.plev[:18])
            T_time = np.arange(np.size(Tglobal.time))
            T_ready = np.flip(Tglobal.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(T_plev, T_time, T_ready)
            Tinterp = interpolated(lev, T_time)

            SH_plev = np.flipud(SHglobal.plev[:18])
            SH_ready = np.flip(SHglobal.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(SH_plev, T_time, SH_ready)
            SHinterp = interpolated(lev, T_time)

            O3_plev = np.flipud(O3global.plev[:18])
            O3_ready = np.flip(O3global.values[:,:18], axis = 1)
            interpolated = interpolate.interp2d(O3_plev, T_time, O3_ready)
            O3interp = interpolated(lev, T_time)

            CL_ready = np.flip(CLglobal.values, axis = 1)
            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CL_ready)
            CLinterp = interpolated(lev, T_time)

            CLW_ready = np.flip(CLWglobal.values, axis = 1)
            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLW_ready)
            CLWinterp = interpolated(lev, T_time)

            CLI_ready = np.flip(CLIglobal.values, axis = 1)
            interpolated = interpolate.interp2d(np.flipud(plev), T_time, CLI_ready)
            CLIinterp = interpolated(lev, T_time)

    #-------------------------- Get top of atmosphere radiative flux data from model for comparison ------------------------#

        #  Incoming SW at TOA
        with xr.open_dataset(constants.variable_to_filename( 'rsdt'), decode_times=True) as rsdt_full:
            rsdt = rsdt_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsdt = (rsdt.rsdt * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at TOA
        with xr.open_dataset(constants.variable_to_filename( 'rsut'), decode_times=True) as rsut_full:
            rsut = rsut_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsut = (rsut.rsut * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at TOA assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rsutcs'), decode_times=True) as rsutcs_full:
            rsutcs = rsutcs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsutcs = (rsutcs.rsutcs * weight).mean(dim=('lat','lon','time'))


        #  Outgoing LW at TOA
        with xr.open_dataset(constants.variable_to_filename( 'rlut'), decode_times=True) as rlut_full:
            rlut = rlut_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlut = (rlut.rlut * weight).mean(dim=('lat','lon','time'))


        #  Outgoing LW at TOA assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rlutcs'), decode_times=True) as rlutcs_full:
            rlutcs = rlutcs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlutcs = (rlutcs.rlutcs * weight).mean(dim=('lat','lon','time'))


        #  Net TOA flux
        if 'MRI' in model:
            rtmt = ( rsdt ) - ( rsut + rlut ) # https://www4.uwsp.edu/geo/faculty/lemke/geog101/lectures/02_radiation_energy_balance.html
        else:
            with xr.open_dataset(constants.variable_to_filename( 'rtmt'), decode_times=True) as rtmt_full:
                rtmt = rtmt_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
                rtmt = (rtmt.rtmt * weight).mean(dim=('lat','lon','time'))


    #-------------------------- Surface radiative flux data --------------------------#

        #  Incoming SW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rsds'), decode_times=True) as rsds_full:
            rsds = rsds_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsds = (rsds.rsds * weight).mean(dim=('lat','lon','time'))


        #  Incoming SW at surface assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rsdscs'), decode_times=True) as rsdscs_full:
            rsdscs = rsdscs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsdscs = (rsdscs.rsdscs * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rsus'), decode_times=True) as rsus_full:
            rsus = rsus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsus = (rsus.rsus * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at surface assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rsuscs'), decode_times=True) as rsuscs_full:
            rsuscs = rsuscs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsuscs = (rsuscs.rsuscs * weight).mean(dim=('lat','lon','time'))


        #  Outgoing LW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rlus'), decode_times=True) as rlus_full:
            rlus = rlus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlus = (rlus.rlus * weight).mean(dim=('lat','lon','time'))


        #  Incoming LW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rlds'), decode_times=True) as rlds_full:
            rlds = rlds_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlds = (rlds.rlds * weight).mean(dim=('lat','lon','time'))


        #  Incoming LW at surface assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rldscs'), decode_times=True) as rldscs_full:
            rldscs = rldscs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rldscs = (rldscs.rldscs * weight).mean(dim=('lat','lon','time'))


        # Determine albedo values at surface and TOA

        albedo_toa = rsut / rsdt
        albedo_toa_cs = rsutcs / rsdt
        albedo_surface = rsus / rsds
        albedo_surface_cs = rsuscs / rsdscs

        model_rad_output = {"rtmt" : rtmt, 
                        "rsdt" : rsdt,
                        "rsut" : rsut,
                        "rsutcs" : rsutcs,
                        "rlut" : rlut,
                        "rlutcs" : rlutcs,
                        "rsds" : rsds,
                        "rsdscs" : rsdscs,
                        "rsus" : rsus,
                        "rsuscs" : rsuscs,
                        "rlus" : rlus,
                        "rlds" : rlds,
                        "rldscs" : rldscs                        
                        }

        #  Setup excel file to store output values
        book = openpyxl.load_workbook( location + 'climate-analysis/reduced_data/model_global_radiation_outputs_dt.xlsx' )
        sheet = book.active

        # Store global means in excel file
        start_row = f
        sheet.cell(row=start_row, column=1).value = name

        for s, val in enumerate(model_rad_output.values()):
            sheet.cell(row=start_row, column=s+2).value = (val.data).item()
        book.save( location + 'climate-analysis/reduced_data//model_global_radiation_outputs_dt.xlsx' )
        f = f+1 # increase row number by 1 to store excel data



        #-------------------------- Convert mixing ratios to water paths --------------------------#


        #  cannot be water content - have tried and gives no effect - values are too small
        air_density = ((lev) / (286.9 * Tinterp))
        clwc = (CLWinterp * air_density) * 1000
        ciwc = (CLIinterp * air_density) * 1000

        prev_alt = 0
        CLWPglobal = np.zeros_like(CLWinterp)
        for i, ( wc, altitude ) in enumerate( zip( np.flipud(clwc), np.flipud(alt) ) ):
            CLWPglobal[i] = wc * ( altitude - prev_alt )
            prev_alt = altitude
        CLWPglobal = np.flipud(CLWPglobal)

        prev_alt = 0
        CIWPglobal = np.zeros_like(CLWinterp)
        for i, ( wc, altitude ) in enumerate( zip( np.flipud(ciwc), np.flipud(alt) ) ):
            CIWPglobal[i] = wc * ( altitude - prev_alt )
            prev_alt = altitude
        CIWPglobal = np.flipud(CIWPglobal)


        #-------------------------- Set dictionaries --------------------------#


        model_CLglobals[name] = CLinterp
        model_CLWPglobals[name] = CLWPglobal
        model_CIWPglobals[name] = CIWPglobal
        model_Tglobals[name] = Tinterp
        model_TSglobals[name] = TSglobal
        model_SHglobals[name] = SHinterp
        model_albedos[name] = albedo_surface
        model_O3globals[name] = O3interp
        model_ch4globals[name] = ch4
        model_n2oglobals[name] = n2o
        model_cfc11globals[name] = cfc11
        model_cfc12globals[name] = cfc12
        model_cfc113globals[name] = cfc113


    #-------------------------- Generate ensemble means --------------------------#

    Tmean = ensemble_mean(model_Tglobals)
    TSmean = ensemble_mean(model_TSglobals)
    CLmean = ensemble_mean(model_CLglobals)
    CLWPmean = ensemble_mean(model_CLWPglobals)
    CIWPmean = ensemble_mean(model_CIWPglobals)
    SHmean = ensemble_mean(model_SHglobals)
    albedomean = ensemble_mean(model_albedos)

    O3mean = ensemble_mean(model_O3globals)
    ch4mean = ensemble_mean(model_ch4globals)
    n2omean = ensemble_mean(model_n2oglobals)
    cfc11mean = ensemble_mean(model_cfc11globals)
    cfc12mean = ensemble_mean(model_cfc12globals)
    cfc113mean = ensemble_mean(model_cfc113globals)

   

    #-------------------------- Generate control data --------------------------#

    #  Set the temperature to the observed values

    control_rsut = 0
    control_rsutcs = 0
    control_rlut = 0
    control_rlutcs = 0
    dt = np.size(Tglobal.time)

    for k in range(0, np.size(Tglobal.time)):

        absorbermean = {'O3': O3mean[k,:], 'CO2': co2, 'CH4':ch4mean, 'N2O':n2omean, 'O2': o2,'CCL4':ccl4, 
            'CFC11':cfc11mean, 'CFC12':cfc12mean, 'CFC113':cfc113mean, 'CFC22':cfc22}

        state.Tatm[:] = Tmean[k,:]
        state.Ts[:] = TSmean[k]

        for i in range(lev.size):
            control_rad = RRTMG(state=state, 
                        albedo=albedomean,
                        absorber_vmr=absorbermean,
                        specific_humidity=SHmean[k,:],
                        cldfrac=CLmean[k,:],
                        verbose=False,
                        clwp = CLWPmean[k,:],
                        ciwp = CIWPmean[k,:],
                        r_liq = np.zeros_like(state.Tatm) + liquid_r,
                        r_ice = np.zeros_like(state.Tatm) + ice_r,               
                        )
            control_rad.compute_diagnostics()
        control_rsut = control_rsut + control_rad.SW_flux_up[0]/dt
        control_rsutcs = control_rsutcs + control_rad.SW_flux_up_clr[0]/dt
        control_rlut = control_rlut + control_rad.LW_flux_up[0]/dt
        control_rlutcs = control_rlutcs + control_rad.LW_flux_up_clr[0]/dt



   
    #  Print and check mean data
    print('===MEAN ENSEMBLE DATA===')
    print('rsut:')
    print(control_rsut)
    print('rsutcs:')
    print(control_rsutcs)
    print('rlut:')
    print(control_rlut)
    print('rlutcs:')
    print(control_rlutcs)


