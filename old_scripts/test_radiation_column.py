import numpy as np
import matplotlib.pyplot as plt
import climlab
from climlab.radiation import RRTMG
import xarray as xr
import constants
import os
from netCDF4 import Dataset
import openpyxl
from cftime import DatetimeNoLeap

location = constants.home
label = 1
model = 'CMIP6-AMIP-CESM2-CAM6'
book = openpyxl.load_workbook( location + 'climate-analysis/reduced_data/radiation_data.xlsx' )
sheet = book.active
lat_bnd_1 = -75
lat_bnd_2 = 75
start = DatetimeNoLeap( 2007, 1, 1 )
end = DatetimeNoLeap( 2010, 1, 1 )

os.chdir( location + 'Data/' + model )


#  Take global, annual average

with xr.open_dataset(constants.variable_to_filename( 'cl' ), decode_times=True) as cl_full:
    cl = cl_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    weight = np.cos(np.deg2rad(cl.lat)) / np.cos(np.deg2rad(cl.lat)).mean(dim='lat')
    CLglobal = (cl.cl * weight).mean(dim=('lat','lon','time')) / 100
    if 'CM4' in model or 'IPSL' in model:
        a = cl.ap
        b = cl.b
    else:
        ap = cl.a
        b = cl.b
        p0 = cl.p0
        a = ap*p0

with xr.open_dataset(constants.variable_to_filename( 'ps' ), decode_times=True) as ps_full:
    ps = ps_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    ps = ps.ps

p = a + b*ps
plev = (p * weight).mean(dim=('lat','lon','time'))

with xr.open_dataset(constants.variable_to_filename( 'clw' ), decode_times=True) as clw_full:
    clw = clw_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    CLWglobal = (clw.clw * weight).mean(dim=('lat','lon','time'))

with xr.open_dataset(constants.variable_to_filename( 'cli' ), decode_times=True) as cli_full:
    cli = cli_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    CLIglobal = (cli.cli * weight).mean(dim=('lat','lon','time'))

with xr.open_dataset(constants.variable_to_filename( 'ta' ), decode_times=True) as ta_full:
    ta = ta_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    Tglobal = (ta.ta * weight).mean(dim=('lat','lon','time'))

with xr.open_dataset(constants.variable_to_filename( 'hus' ), decode_times=True) as hus_full:
    hus = hus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    SHglobal = (hus.hus * weight).mean(dim=('lat','lon','time'))  # kg/kg

with xr.open_dataset(constants.variable_to_filename( 'o3' ), decode_times=True) as o3_full:
    o3 = o3_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    O3global = (o3.o3 * weight).mean(dim=('lat','lon','time'))  # kg/kg

with xr.open_dataset(constants.variable_to_filename( 'rsus' ), decode_times=True) as rsus_full:
    rsus = rsus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    RSUSglobal = (rsus.rsus * weight).mean(dim=('lat','lon','time'))  # kg/kg

with xr.open_dataset(constants.variable_to_filename( 'rsds' ), decode_times=True) as rsds_full:
    rsds = rsds_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    RSDSglobal = (rsds.rsds * weight).mean(dim=('lat','lon','time'))  # kg/kg

with xr.open_dataset(constants.variable_to_filename( 'ts' ), decode_times=True) as ts_full:
    ts = ts_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
    TSglobal = (ts.ts * weight).mean(dim=('lat','lon','time'))  # kg/kg

albedo_surface = RSUSglobal / RSDSglobal


state = climlab.column_state(num_lev=50)
lev = state.Tatm.domain.axes['lev'].points*100

#  interpolate to model pressure levels
#  Need to 'flipud' because the interpolation routine 
#  needs the pressure data to be in increasing order
#  Create a state dictionary with corresponsing number of cl data levels

if 'CAM6' in model:
    Tinterp = np.interp(lev, np.flipud(Tglobal.plev[:18]), np.flipud(Tglobal[:18]))
    SHinterp = np.interp(lev, np.flipud(SHglobal.plev[:18]), np.flipud(SHglobal[:18]))
    O3interp = np.interp(lev, np.flipud(O3global.plev[:18]), np.flipud(O3global[:18]))
    CLinterp = np.interp(lev, (plev), (CLglobal))
    CLWinterp = np.interp(lev, (plev), (CLWglobal))
    CLIinterp = np.interp(lev, (plev), (CLIglobal))
else:

    Tinterp = np.interp(lev, np.flipud(Tglobal.plev[:18]), np.flipud(Tglobal[:18]))
    SHinterp = np.interp(lev, np.flipud(SHglobal.plev[:18]), np.flipud(SHglobal[:18]))
    O3interp = np.interp(lev, np.flipud(O3global.plev[:18]), np.flipud(O3global[:18]))
    CLinterp = np.interp(lev, np.flipud(plev), np.flipud(CLglobal))
    CLWinterp = np.interp(lev, np.flipud(plev), np.flipud(CLWglobal))
    CLIinterp = np.interp(lev, np.flipud(plev), np.flipud(CLIglobal))

#  Set the temperature to the observed values
state.Tatm[:] = Tinterp
state.Ts[:] = TSglobal

r_liq = 14.  # Cloud water drop effective radius (microns)
r_ice = 60.  # Cloud water drop effective radius (microns)

#  Convert mixing ratio (kg/kg) into cloud water content in (g/m3)

alt = np.flipud(constants.p_to_alt( np.flipud(lev) )  * 1000) # in m
air_density = ((lev) / (286.9 * Tinterp))
clwc = (CLWinterp * air_density) * 1000
ciwc = (CLIinterp * air_density) * 1000

prev_alt = alt[0] + (alt[0] - alt[1])
CLWPglobal = np.zeros_like(lev)
for i, ( row, altitude ) in enumerate( zip( clwc, alt ) ):
    CLWPglobal[i] = row * (prev_alt - altitude)
    prev_alt = altitude

prev_alt = alt[-1] + (alt[-1] - alt[-2])
CIWPglobal = np.zeros_like(lev)
for i, ( row, altitude ) in enumerate( zip( ciwc, alt ) ):
    CIWPglobal[i] = row * (prev_alt - altitude)
    prev_alt = altitude
#  Convert cloud water content to water path (g/m2)
#  WP = WC * deltaz
# CLWPglobal = constants.wc_to_wp( clwc, alt )
# CIWPglobal = constants.wc_to_wp( ciwc, alt )


#  Loop through all pressure levels
#  Set up a radiation model with the cloud layer at the current pressure level
#  Compute CRE and store the results
absorber = {'O3': O3interp, 'CO2': 400e-6, 'CH4':1.557889297535934e-6, 'N2O':3.007494866529774e-7, 'O2': 0.21,'CCL4':0., 
    'CFC11':1.680e-10, 'CFC12':2.850e-10, 'CFC113':1.737268993839836e-11, 'CFC22':4.8743326488706363e-11}

CRE_LW = {}
CRE_SW = {}
OLR = np.zeros_like(lev)
ASR = np.zeros_like(lev)
OLRclr = np.zeros_like(lev)
ASRclr = np.zeros_like(lev)
for i in range(lev.size):
    # Whole-column cloud characteristics
    # The cloud fraction is a Gaussian bump centered at the current level        
    rad = RRTMG(state=state, 
                albedo=albedo_surface,
                absorber_vmr=absorber,
                specific_humidity=SHinterp,
                cldfrac=CLinterp*np.exp(-(lev-lev[i])**2/(2*25.)**2),
                verbose=False,
                clwp = CLWPglobal*1000,
                ciwp = CIWPglobal*1000,
                r_liq = np.zeros_like(state.Tatm) + r_liq,
                r_ice = np.zeros_like(state.Tatm) + r_ice,               
                )
    rad.compute_diagnostics()
    OLR[i] = rad.OLR
    OLRclr[i] = rad.OLRclr
    ASR[i] = rad.ASR
    ASRclr[i] = rad.ASRclr
   
CRE_LW = -(OLR - OLRclr)
CRE_SW = (ASR - ASRclr)
 
#  Make some plots of the CRE dependence on cloud height
fig, axes = plt.subplots(1,3, figsize=(16,6))
ax = axes[0]
ax.plot(rad.SW_flux_net[1:], alt/1000)
ax.set_ylabel('Altitude (km)')
ax.set_xlabel('SW cloud radiative effect (W/m2)')

ax = axes[1]
ax.plot(rad.LW_flux_net[1:], alt/1000)
ax.set_xlabel('LW cloud radiative effect (W/m2)')

ax = axes[2]
ax.plot(CRE_SW + CRE_LW, alt/1000)
ax.set_xlabel('Net cloud radiative effect (W/m2)')

for ax in axes:
    # ax.invert_yaxis()
    # ax.legend()
    ax.grid()
fig.suptitle('Cloud Radiative Effect as a function of the vertical height of the cloud layer', fontsize=16)
plt.show()

# Store global means in excel file
start_row = int(label) + 2
sheet.cell(row=start_row, column=1).value = label
sheet.cell(row=start_row, column=2).value = model
sheet.cell(row=start_row, column=3).value = r_liq
sheet.cell(row=start_row, column=4).value = r_ice
sheet.cell(row=start_row, column=5).value = rad.SW_flux_net[0]
sheet.cell(row=start_row, column=6).value = rad.LW_flux_net[0]
sheet.cell(row=start_row, column=7).value = np.mean(CRE_SW)
sheet.cell(row=start_row, column=8).value = np.mean(CRE_LW)
sheet.cell(row=start_row, column=9).value = rad.SW_flux_net[0] - rad.LW_flux_net[0]
book.save( location + 'climate-analysis/reduced_data/radiation_data.xlsx' )


with xr.open_dataset(constants.variable_to_filename( 'rsutcs' ), decode_times=False) as rsutcs_full:
    rsutcs = rsutcs_full.sel(lat=slice(lat_bnd_1, lat_bnd_2))
    model_sw_up_cs = (rsutcs.rsutcs * weight).mean(dim=('lat','lon','time'))

with xr.open_dataset(constants.variable_to_filename( 'rlutcs' ), decode_times=False) as rlutcs_full:
    rlutcs = rlutcs_full.sel(lat=slice(lat_bnd_1, lat_bnd_2))
    model_lw_up_cs = (rlutcs.rlutcs * weight).mean(dim=('lat','lon','time'))

print (model_sw_up_cs.values)
print (rad.SW_flux_up_clr[0])
print (model_lw_up_cs.values)
print (rad.LW_flux_up_clr[0])
