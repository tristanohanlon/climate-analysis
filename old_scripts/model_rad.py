

import numpy as np
import os
from netCDF4 import Dataset
from netCDF4 import date2index
import h5py
import math
from scipy import interpolate
from scipy.interpolate import RectBivariateSpline
from scipy import stats
import datetime
from pprint import pprint
from sklearn.impute import SimpleImputer
import constants
from scipy import ndimage as nd
import matplotlib.pyplot as plt
import climlab
from climlab.radiation import RRTMG
from matplotlib import rc
rc('text', usetex=True)
from matplotlib.font_manager import FontProperties
fontP = FontProperties()
fontP.set_size('small')
import openpyxl
import xarray as xr



def ensemble_mean( ref_dict ):
    running_sum = np.zeros_like(next(iter(ref_dict.values())))
    for items in ref_dict.values():
        running_sum += items
    mean = running_sum / len(ref_dict)
    return mean


def radiation(start, end, start_dt, end_dt, location, models, label, lat_bnd_1, lat_bnd_2, save_outputs, liquid_r, ice_r ):



    #  Set aerosol and level constants
    o2=0.21
    ccl4=0.
    cfc22=4.8743326488706363e-11
    co2 = 400.e-6

    state = climlab.column_state(num_lev=50)
    lev = state.Tatm.domain.axes['lev'].points*100
    alt = np.flipud(constants.p_to_alt( np.flipud(lev) )  * 1000) # in m
    f = 2 # counter for excel rows
    if label == '2':
        f = 10

    model_CLglobals = {}
    model_CLWPglobals = {}
    model_CIWPglobals = {}
    model_Tglobals = {}
    model_TSglobals = {}
    model_SHglobals = {}
    model_albedos = {}

    model_O3globals = {}
    model_ch4globals = {}
    model_n2oglobals = {}
    model_cfc11globals = {}
    model_cfc12globals = {}
    model_cfc113globals = {}
    model_INSOLATIONglobals = {}
    
    SWcre_CLdeltas = {}
    LWcre_CLdeltas = {}

    SWcre_LWPdeltas = {}
    LWcre_LWPdeltas = {}

    SWcre_IWPdeltas = {}
    LWcre_IWPdeltas = {}

    SWcre_Tdeltas = {}
    LWcre_Tdeltas = {}

    SWcre_SHdeltas = {}
    LWcre_SHdeltas = {}


    SWcs_CLdeltas = {}
    LWcs_CLdeltas = {}

    SWcs_LWPdeltas = {}
    LWcs_LWPdeltas = {}

    SWcs_IWPdeltas = {}
    LWcs_IWPdeltas = {}

    SWcs_Tdeltas = {}
    LWcs_Tdeltas = {}

    SWcs_SHdeltas = {}
    LWcs_SHdeltas = {}


    SWcre_lr5deltas = {}
    LWcre_lr5deltas = {}

    SWcre_lr25deltas = {}
    LWcre_lr25deltas = {}

    SWcre_lr60deltas = {}
    LWcre_lr60deltas = {}

    SWcre_ir30deltas = {}
    LWcre_ir30deltas = {}

    SWcre_ir80deltas = {}
    LWcre_ir80deltas = {}

    SWcre_ir130deltas = {}
    LWcre_ir130deltas = {}



    colors = ["crimson", "purple", "limegreen", "gold", "blue", "magenta"]

    #  Get global mean valriables for all models and interpolate to common levels
    for name, model in models.items():
        os.chdir( location + 'Data/' + model )

        if 'IPSL' in name or 'MRI' in name or 'MIROC' in name:
            start = start_dt
            end = end_dt

    #----------------------------- Cloud variables -----------------------------#

        #  Take global, annual average

        with xr.open_dataset(constants.variable_to_filename( 'cl' ), decode_times=True) as cl_full:
            cl = cl_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
        weight = np.cos(np.deg2rad(cl.lat)) / np.cos(np.deg2rad(cl.lat)).mean(dim='lat')
        CLglobal = (cl.cl * weight).mean(dim=('lat','lon','time')) / 100

        if 'CM4' in model or 'IPSL' in model:
            a = cl.ap
            b = cl.b
        else:
            ap = cl.a
            b = cl.b
            p0 = cl.p0
            a = ap*p0

        with xr.open_dataset(constants.variable_to_filename( 'ps' ), decode_times=True) as ps_full:
            ps = ps_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            ps = ps.ps

        p = a + b*ps
        plev = (p * weight).mean(dim=('lat','lon','time'))
        if 'IPSL' in model:
            plev = plev[:-1]


        with xr.open_dataset(constants.variable_to_filename( 'clw' ), decode_times=True) as clw_full:
            clw = clw_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            CLWglobal = (clw.clw * weight).mean(dim=('lat','lon','time'))

        with xr.open_dataset(constants.variable_to_filename( 'cli' ), decode_times=True) as cli_full:
            cli = cli_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            CLIglobal = (cli.cli * weight).mean(dim=('lat','lon','time'))

        with xr.open_dataset(constants.variable_to_filename( 'ta' ), decode_times=True) as ta_full:
            ta = ta_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            Tglobal = (ta.ta * weight).mean(dim=('lat','lon','time'))

        with xr.open_dataset(constants.variable_to_filename( 'hus' ), decode_times=True) as hus_full:
            hus = hus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            SHglobal = (hus.hus * weight).mean(dim=('lat','lon','time'))  # kg/kg

        with xr.open_dataset(constants.variable_to_filename( 'ts' ), decode_times=True) as ts_full:
            ts = ts_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            TSglobal = (ts.ts * weight).mean(dim=('lat','lon','time'))  # kg/kg



                #----------------------------- Aerosols ---------------------------#

        #  Get aerosol values from model files or set default values if they are
        #  not present.
        #  Convert mol/mol to ppm by dividing by 1E-6
        #  Convery CO2 mass to ppm by dividing by atmosphere mass 4.99E18
    
        #  Mole Fraction of Ozone (time, lev, lat, lon)

        check_file = constants.variable_to_filename( 'o3' )
        if check_file == None:
            # use current global mean
            O3global = ensemble_mean(model_O3globals)

        else:
            with xr.open_dataset(constants.variable_to_filename( 'o3'), decode_times=True) as o3_full:
                o3 = o3_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
                O3global = (o3.o3 * weight).mean(dim=('lat','lon','time'))  # kg/kg


        #  Global Mean Mole Fraction of CH4 (time)

        check_file = constants.variable_to_filename( 'ch4global' )
        if check_file == None:
            ch4 = 1.557889297535934e-6
        else:
            with xr.open_dataset(constants.variable_to_filename( 'ch4global' ), decode_times=True) as ch4_full:
                ch4 = ch4_full.sel(time=slice(start,end))
                ch4 = (ch4.ch4global).mean(dim=('time'))
                ch4 = ch4.values / 1e9


        #  Global Mean Mole Fraction of N2O (time)

        check_file = constants.variable_to_filename( 'n2oglobal' )
        if check_file == None:
            n2o = 3.007494866529774e-7
        else:
            with xr.open_dataset(constants.variable_to_filename( 'n2oglobal' ), decode_times=True) as n2o_full:
                n2o = n2o_full.sel(time=slice(start,end))
                n2o = (n2o.n2oglobal).mean(dim=('time'))
                n2o = n2o.values / 1e9


        #  Global Mean Mole Fraction of CFC11 (time)

        check_file = constants.variable_to_filename( 'cfc11global' )
        if check_file == None:
            cfc11 = 1.680e-10
        else:
            with xr.open_dataset(constants.variable_to_filename( 'cfc11global' ), decode_times=True) as cfc11_full:
                cfc11 = cfc11_full.sel(time=slice(start,end))
                cfc11 = (cfc11.cfc11global).mean(dim=('time'))
                cfc11 = cfc11.values / 1e12


        #  Global Mean Mole Fraction of CFC12 (time)

        check_file = constants.variable_to_filename( 'cfc12global' )
        if check_file == None:
            cfc12 = 2.850e-10
        else:
            with xr.open_dataset(constants.variable_to_filename( 'cfc12global' ), decode_times=True) as cfc12_full:
                cfc12 = cfc12_full.sel(time=slice(start,end))
                cfc12 = (cfc12.cfc12global).mean(dim=('time'))
                cfc12 = cfc12.values / 1e12


        #  Global Mean Mole Fraction of CFC113 (time)

        check_file = constants.variable_to_filename( 'cfc113global' )
        if check_file == None:
            cfc113 = 1.737268993839836e-11
        else:
            with xr.open_dataset(constants.variable_to_filename( 'cfc113global' ), decode_times=True) as cfc113_full:
                cfc113 = cfc113_full.sel(time=slice(start,end))
                cfc113 = (cfc113.cfc113global).mean(dim=('time'))
                cfc113 = cfc113.values / 1e12




        #  interpolate to model pressure levels
        #  Need to 'flipud' because the interpolation routine 
        #  needs the pressure data to be in increasing order
        #  Create a state dictionary with corresponsing number of cl data levels

        if 'CAM6' in model:
            Tinterp = np.interp(lev, np.flipud(Tglobal.plev[:18]), np.flipud(Tglobal[:18]))
            SHinterp = np.interp(lev, np.flipud(SHglobal.plev[:18]), np.flipud(SHglobal[:18]))
            O3interp = np.interp(lev, np.flipud(O3global.plev[:18]), np.flipud(O3global[:18]))
            CLinterp = np.interp(lev, (plev), (CLglobal))
            CLWinterp = np.interp(lev, (plev), (CLWglobal))
            CLIinterp = np.interp(lev, (plev), (CLIglobal))

        elif 'MRI' in model or 'MIROC' in model:
            Tinterp = np.interp(lev, np.flipud(Tglobal.plev[:18]), np.flipud(Tglobal[:18]))
            SHinterp = np.interp(lev, np.flipud(SHglobal.plev[:18]), np.flipud(SHglobal[:18]))
            O3interp = O3global
            CLinterp = np.interp(lev, np.flipud(plev), np.flipud(CLglobal))
            CLWinterp = np.interp(lev, np.flipud(plev), np.flipud(CLWglobal))
            CLIinterp = np.interp(lev, np.flipud(plev), np.flipud(CLIglobal))

        else:

            Tinterp = np.interp(lev, np.flipud(Tglobal.plev[:18]), np.flipud(Tglobal[:18]))
            SHinterp = np.interp(lev, np.flipud(SHglobal.plev[:18]), np.flipud(SHglobal[:18]))
            O3interp = np.interp(lev, np.flipud(O3global.plev[:18]), np.flipud(O3global[:18]))
            CLinterp = np.interp(lev, np.flipud(plev), np.flipud(CLglobal))
            CLWinterp = np.interp(lev, np.flipud(plev), np.flipud(CLWglobal))
            CLIinterp = np.interp(lev, np.flipud(plev), np.flipud(CLIglobal))


    #-------------------------- Get top of atmosphere radiative flux data from model for comparison ------------------------#

        #  Incoming SW at TOA
        with xr.open_dataset(constants.variable_to_filename( 'rsdt'), decode_times=True) as rsdt_full:
            rsdt = rsdt_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsdt = (rsdt.rsdt * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at TOA
        with xr.open_dataset(constants.variable_to_filename( 'rsut'), decode_times=True) as rsut_full:
            rsut = rsut_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsut = (rsut.rsut * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at TOA assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rsutcs'), decode_times=True) as rsutcs_full:
            rsutcs = rsutcs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsutcs = (rsutcs.rsutcs * weight).mean(dim=('lat','lon','time'))


        #  Outgoing LW at TOA
        with xr.open_dataset(constants.variable_to_filename( 'rlut'), decode_times=True) as rlut_full:
            rlut = rlut_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlut = (rlut.rlut * weight).mean(dim=('lat','lon','time'))


        #  Outgoing LW at TOA assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rlutcs'), decode_times=True) as rlutcs_full:
            rlutcs = rlutcs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlutcs = (rlutcs.rlutcs * weight).mean(dim=('lat','lon','time'))


        #  Net TOA flux
        if 'MRI' in model:
            rtmt = ( rsdt ) - ( rsut + rlut ) # https://www4.uwsp.edu/geo/faculty/lemke/geog101/lectures/02_radiation_energy_balance.html
        else:
            with xr.open_dataset(constants.variable_to_filename( 'rtmt'), decode_times=True) as rtmt_full:
                rtmt = rtmt_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
                rtmt = (rtmt.rtmt * weight).mean(dim=('lat','lon','time'))


    #-------------------------- Surface radiative flux data --------------------------#

        #  Incoming SW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rsds'), decode_times=True) as rsds_full:
            rsds = rsds_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsds = (rsds.rsds * weight).mean(dim=('lat','lon','time'))


        #  Incoming SW at surface assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rsdscs'), decode_times=True) as rsdscs_full:
            rsdscs = rsdscs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsdscs = (rsdscs.rsdscs * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rsus'), decode_times=True) as rsus_full:
            rsus = rsus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsus = (rsus.rsus * weight).mean(dim=('lat','lon','time'))


        #  Outgoing SW at surface assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rsuscs'), decode_times=True) as rsuscs_full:
            rsuscs = rsuscs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rsuscs = (rsuscs.rsuscs * weight).mean(dim=('lat','lon','time'))


        #  Outgoing LW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rlus'), decode_times=True) as rlus_full:
            rlus = rlus_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlus = (rlus.rlus * weight).mean(dim=('lat','lon','time'))


        #  Incoming LW at surface
        with xr.open_dataset(constants.variable_to_filename( 'rlds'), decode_times=True) as rlds_full:
            rlds = rlds_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rlds = (rlds.rlds * weight).mean(dim=('lat','lon','time'))


        #  Incoming LW at surface assuming clear sky
        with xr.open_dataset(constants.variable_to_filename( 'rldscs'), decode_times=True) as rldscs_full:
            rldscs = rldscs_full.sel(time=slice(start,end), lat=slice(lat_bnd_1, lat_bnd_2))
            rldscs = (rldscs.rldscs * weight).mean(dim=('lat','lon','time'))


        # Determine albedo values at surface and TOA

        albedo_toa = rsut / rsdt
        albedo_toa_cs = rsutcs / rsdt
        albedo_surface = rsus / rsds
        albedo_surface_cs = rsuscs / rsdscs

        model_rad_output = {"rtmt" : rtmt, 
                        "rsdt" : rsdt,
                        "rsut" : rsut,
                        "rsutcs" : rsutcs,
                        "rlut" : rlut,
                        "rlutcs" : rlutcs,
                        "rsds" : rsds,
                        "rsdscs" : rsdscs,
                        "rsus" : rsus,
                        "rsuscs" : rsuscs,
                        "rlus" : rlus,
                        "rlds" : rlds,
                        "rldscs" : rldscs                        
                        }

        #  Setup excel file to store output values
        # book = openpyxl.load_workbook( location + 'climate-analysis/reduced_data/model_global_radiation_outputs.xlsx' )
        # sheet = book.active

        # # Store global means in excel file
        # start_row = f
        # sheet.cell(row=start_row, column=1).value = name

        # for s, val in enumerate(model_rad_output.values()):
        #     sheet.cell(row=start_row, column=s+2).value = (val.data).item()
        # book.save( location + 'climate-analysis/reduced_data//model_global_radiation_outputs.xlsx' )
        # f = f+1 # increase row number by 1 to store excel data



        #-------------------------- Convert mixing ratios to water paths --------------------------#


        #  cannot be water content - have tried and gives no effect - values are too small
        air_density = ((lev) / (286.9 * Tinterp))
        clwc = (CLWinterp * air_density) * 1000
        ciwc = (CLIinterp * air_density) * 1000

        prev_alt = 0
        CLWPglobal = np.zeros_like(lev)
        for i, ( wc, altitude ) in enumerate( zip( np.flipud(clwc), np.flipud(alt) ) ):
            CLWPglobal[i] = wc * ( altitude - prev_alt )
            prev_alt = altitude
        CLWPglobal = np.flipud(CLWPglobal)

        prev_alt = 0
        CIWPglobal = np.zeros_like(lev)
        for i, ( wc, altitude ) in enumerate( zip( np.flipud(ciwc), np.flipud(alt) ) ):
            CIWPglobal[i] = wc * ( altitude - prev_alt )
            prev_alt = altitude
        CIWPglobal = np.flipud(CIWPglobal)


        #-------------------------- Set dictionaries --------------------------#


        model_CLglobals[name] = CLinterp
        model_CLWPglobals[name] = CLWPglobal
        model_CIWPglobals[name] = CIWPglobal
        model_Tglobals[name] = Tinterp
        model_TSglobals[name] = TSglobal
        model_SHglobals[name] = SHinterp
        model_albedos[name] = albedo_surface
        model_O3globals[name] = O3interp
        model_ch4globals[name] = ch4
        model_n2oglobals[name] = n2o
        model_cfc11globals[name] = cfc11
        model_cfc12globals[name] = cfc12
        model_cfc113globals[name] = cfc113
        model_INSOLATIONglobals[name] = rsdt


    #-------------------------- Generate ensemble means --------------------------#

    Tmean = ensemble_mean(model_Tglobals)
    TSmean = ensemble_mean(model_TSglobals)
    CLmean = ensemble_mean(model_CLglobals)
    CLWPmean = ensemble_mean(model_CLWPglobals)
    CIWPmean = ensemble_mean(model_CIWPglobals)
    SHmean = ensemble_mean(model_SHglobals)
    albedomean = ensemble_mean(model_albedos)

    O3mean = ensemble_mean(model_O3globals)
    ch4mean = ensemble_mean(model_ch4globals)
    n2omean = ensemble_mean(model_n2oglobals)
    cfc11mean = ensemble_mean(model_cfc11globals)
    cfc12mean = ensemble_mean(model_cfc12globals)
    cfc113mean = ensemble_mean(model_cfc113globals)
    INSOLATIONmean = ensemble_mean(model_INSOLATIONglobals)

    absorbermean = {'O3': O3mean, 'CO2': co2, 'CH4':ch4mean, 'N2O':n2omean, 'O2': o2,'CCL4':ccl4, 
        'CFC11':cfc11mean, 'CFC12':cfc12mean, 'CFC113':cfc113mean, 'CFC22':cfc22}


    #-------------------------- Check ozone profiles --------------------------#


    # fig, ax = plt.subplots()
    # fig.suptitle('Ozone Profiles')

    # for z, (name, data) in enumerate(model_O3globals.items()):
    #     ax.plot(data, alt, color=colors[z], label=name)
   

    # ax.legend(bbox_to_anchor=(1, 1))
    # ax.set_ylabel('Altitude $km$')
    # ax.set_xlabel('$O_3$ ($mol/mol$)')
    # plt.savefig( location + 'Images/RRTGM/' + label + "_o3.pdf", format="pdf", bbox_inches='tight')
    # plt.show()


    

    #-------------------------- Generate control data --------------------------#

    #  Set the temperature to the observed values
    state.Tatm[:] = Tmean
    state.Ts[:] = TSmean

    for i in range(lev.size):
        control_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    insolation = INSOLATIONmean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + liquid_r,
                    r_ice = np.zeros_like(state.Tatm) + ice_r,               
                    )
        control_rad.compute_diagnostics()
   
    #  Print and check mean data
    print('===MEAN ENSEMBLE DATA===')
    print('rsut:')
    print(control_rad.SW_flux_up[0])
    print('rsutcs:')
    print(control_rad.SW_flux_up_clr[0])
    print('rsds:')
    print(control_rad.SW_flux_down[-1])
    print('rsdscs:')
    print(control_rad.SW_flux_down_clr[-1])
    print('rlut:')
    print(control_rad.LW_flux_up[0])
    print('rlutcs:')
    print(control_rad.LW_flux_up_clr[0])
    print('rldscs:')
    print(control_rad.LW_flux_down_clr[-1])

    labels = ["rsutcs", "rsdscs", "rlutcs"]

    rad_data = [control_rad.SW_flux_up_clr[0], control_rad.SW_flux_down_clr[-1], control_rad.LW_flux_up_clr[0]]
    model_data = [51.5, 245.4, 263.8]
    width=0.4
    fig, ax = plt.subplots()
    
    label_added = False
    for i, data in enumerate(rad_data):
        if not label_added:
            ax.scatter(i, data, s=25, marker=".", label="RRTMG Output", c="k")
            label_added = True
        else:
            ax.scatter(i, data, s=25, marker=".", c="k")
    
    label_added = False
    for i, data in enumerate(model_data):
        if not label_added:
            ax.scatter(i, data, s=25, marker="x", label="Model Output", c="k")
            label_added = True
        else:
            ax.scatter(i, data, s=25, marker="x", c="k")


    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels)
    ax.legend(bbox_to_anchor=(1, 1))
    ax.set_ylabel('$Wm^{-2}$')
    plt.savefig( location + 'Images/RRTGM/' + label + "_model_rad_output_compare.pdf", format="pdf", bbox_inches='tight')
    plt.show()



    #-------------------------- Generate CL Bias Data --------------------------#

    for model_name, model_cl_value in model_CLglobals.items():
        for i in range(lev.size):
            cl_rad = RRTMG(state=state, 
                        albedo=albedomean,
                        absorber_vmr=absorbermean,
                        specific_humidity=SHmean,
                        cldfrac=model_cl_value, # compute variances from these model values
                        verbose=False,
                        clwp = CLWPmean,
                        ciwp = CIWPmean,
                        r_liq = np.zeros_like(state.Tatm) + liquid_r,
                        r_ice = np.zeros_like(state.Tatm) + ice_r,               
                        )
            cl_rad.compute_diagnostics()
        SWcre_CLdeltas[model_name] = (cl_rad.SW_flux_up[0] - cl_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
        LWcre_CLdeltas[model_name] = (cl_rad.LW_flux_up_clr[0] - cl_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])

        SWcs_CLdeltas[model_name] = (cl_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up_clr[0])
        LWcs_CLdeltas[model_name] = (cl_rad.LW_flux_up_clr[0]) - (control_rad.LW_flux_up_clr[0])

    #-------------------------- Generate CLWP Bias Data --------------------------#

    for model_name, model_clwp_value in model_CLWPglobals.items():
        for i in range(lev.size):
            clwp_rad = RRTMG(state=state, 
                        albedo=albedomean,
                        absorber_vmr=absorbermean,
                        specific_humidity=SHmean,
                        cldfrac=CLmean,
                        verbose=False,
                        clwp = model_clwp_value, # compute variances from these model values
                        ciwp = CIWPmean,
                        r_liq = np.zeros_like(state.Tatm) + liquid_r,
                        r_ice = np.zeros_like(state.Tatm) + ice_r,               
                        )
            clwp_rad.compute_diagnostics()
        SWcre_LWPdeltas[model_name] = (clwp_rad.SW_flux_up[0] - clwp_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
        LWcre_LWPdeltas[model_name] = (clwp_rad.LW_flux_up_clr[0] - clwp_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])

        SWcs_LWPdeltas[model_name] = (clwp_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up_clr[0])
        LWcs_LWPdeltas[model_name] = (clwp_rad.LW_flux_up_clr[0]) - (control_rad.LW_flux_up_clr[0])


    #-------------------------- Generate CIWP Bias Data --------------------------#

    for model_name, model_ciwp_value in model_CIWPglobals.items():
        for i in range(lev.size):
            ciwp_rad = RRTMG(state=state, 
                        albedo=albedomean,
                        absorber_vmr=absorbermean,
                        specific_humidity=SHmean,
                        cldfrac=CLmean,
                        verbose=False,
                        clwp = CLWPmean, 
                        ciwp = model_ciwp_value, # compute variances from these model values
                        r_liq = np.zeros_like(state.Tatm) + liquid_r,
                        r_ice = np.zeros_like(state.Tatm) + ice_r,               
                        )
            ciwp_rad.compute_diagnostics()
        SWcre_IWPdeltas[model_name] = (ciwp_rad.SW_flux_up[0] - ciwp_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
        LWcre_IWPdeltas[model_name] = (ciwp_rad.LW_flux_up_clr[0] - ciwp_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])

        SWcs_IWPdeltas[model_name] = (ciwp_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up_clr[0])
        LWcs_IWPdeltas[model_name] = (ciwp_rad.LW_flux_up_clr[0]) - (control_rad.LW_flux_up_clr[0])


    #-------------------------- Generate T Bias Data --------------------------#

    for model_name, model_T_value in model_Tglobals.items():
        state.Tatm[:] = model_T_value  # compute variances from these model values
        for i in range(lev.size):
            T_rad = RRTMG(state=state, 
                        albedo=albedomean,
                        absorber_vmr=absorbermean,
                        specific_humidity=SHmean,
                        cldfrac=CLmean,
                        verbose=False,
                        clwp = CLWPmean,
                        ciwp = CIWPmean,
                        r_liq = np.zeros_like(state.Tatm) + liquid_r,
                        r_ice = np.zeros_like(state.Tatm) + ice_r,               
                        )
            T_rad.compute_diagnostics()
        SWcre_Tdeltas[model_name] = (T_rad.SW_flux_up[0] - T_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
        LWcre_Tdeltas[model_name] = (T_rad.LW_flux_up_clr[0] - T_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])

        SWcs_Tdeltas[model_name] = (T_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up_clr[0])
        LWcs_Tdeltas[model_name] = (T_rad.LW_flux_up_clr[0]) - (control_rad.LW_flux_up_clr[0])
    state.Tatm[:] = Tmean


    #-------------------------- Generate SH Bias Data --------------------------#

    for model_name, model_SH_value in model_SHglobals.items():
        for i in range(lev.size):
            SH_rad = RRTMG(state=state, 
                        albedo=albedomean,
                        absorber_vmr=absorbermean,
                        specific_humidity=model_SH_value, # compute variances from these model values
                        cldfrac=CLmean,
                        verbose=False,
                        clwp = CLWPmean,
                        ciwp = CIWPmean,
                        r_liq = np.zeros_like(state.Tatm) + liquid_r,
                        r_ice = np.zeros_like(state.Tatm) + ice_r,               
                        )
            SH_rad.compute_diagnostics()
        SWcre_SHdeltas[model_name] = (SH_rad.SW_flux_up[0] - SH_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
        LWcre_SHdeltas[model_name] = (SH_rad.LW_flux_up_clr[0] - SH_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])

        SWcs_SHdeltas[model_name] = (SH_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up_clr[0])
        LWcs_SHdeltas[model_name] = (SH_rad.LW_flux_up_clr[0]) - (control_rad.LW_flux_up_clr[0])



    #-------------------------- Generate Liquid Particle Radius Change Data --------------------------#


    for i in range(lev.size):
        lr5_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + 15.0,  # compute variances from these model values
                    r_ice = np.zeros_like(state.Tatm) + ice_r,               
                    )
        lr5_rad.compute_diagnostics()
    SWcre_lr5deltas = (lr5_rad.SW_flux_up[0] - lr5_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
    LWcre_lr5deltas = (lr5_rad.LW_flux_up_clr[0] - lr5_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])


    for i in range(lev.size):
        lr25_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + 45.0,  # compute variances from these model values
                    r_ice = np.zeros_like(state.Tatm) + ice_r,               
                    )
        lr25_rad.compute_diagnostics()
    SWcre_lr25deltas = (lr25_rad.SW_flux_up[0] - lr25_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
    LWcre_lr25deltas = (lr25_rad.LW_flux_up_clr[0] - lr25_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])


    for i in range(lev.size):
        lr60_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + 60.0,  # compute variances from these model values
                    r_ice = np.zeros_like(state.Tatm) + ice_r,               
                    )
        lr60_rad.compute_diagnostics()
    SWcre_lr60deltas = (lr60_rad.SW_flux_up[0] - lr60_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
    LWcre_lr60deltas = (lr60_rad.LW_flux_up_clr[0] - lr60_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])


    #-------------------------- Generate Ice Particle Radius Change Data --------------------------#


    for i in range(lev.size):
        ir30_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + liquid_r,  # compute variances from these model values
                    r_ice = np.zeros_like(state.Tatm) + 30.0,               
                    )
        ir30_rad.compute_diagnostics()
    SWcre_ir30deltas = (ir30_rad.SW_flux_up[0] - ir30_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
    LWcre_ir30deltas = (ir30_rad.LW_flux_up_clr[0] - ir30_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])


    for i in range(lev.size):
        ir80_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + liquid_r,  # compute variances from these model values
                    r_ice = np.zeros_like(state.Tatm) + 80.0,               
                    )
        ir80_rad.compute_diagnostics()
    SWcre_ir80deltas = (ir80_rad.SW_flux_up[0] - ir80_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
    LWcre_ir80deltas = (ir80_rad.LW_flux_up_clr[0] - ir80_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])


    for i in range(lev.size):
        ir130_rad = RRTMG(state=state, 
                    albedo=albedomean,
                    absorber_vmr=absorbermean,
                    specific_humidity=SHmean,
                    cldfrac=CLmean,
                    verbose=False,
                    clwp = CLWPmean,
                    ciwp = CIWPmean,
                    r_liq = np.zeros_like(state.Tatm) + liquid_r,  # compute variances from these model values
                    r_ice = np.zeros_like(state.Tatm) + 130.0,               
                    )
        ir130_rad.compute_diagnostics()
    SWcre_ir130deltas = (ir130_rad.SW_flux_up[0] - ir130_rad.SW_flux_up_clr[0]) - (control_rad.SW_flux_up[0] - control_rad.SW_flux_up_clr[0])
    LWcre_ir130deltas = (ir130_rad.LW_flux_up_clr[0] - ir130_rad.LW_flux_up[0]) - (control_rad.LW_flux_up_clr[0] - control_rad.LW_flux_up[0])


 #-------------------------- Build data plots --------------------------#


    #  CRE Plots
    
    labels = ["$\Delta cl$", "$\Delta lwp$", "$\Delta iwp$", "$\Delta T$", "$\Delta SH$"]
    sw_cre_data = [SWcre_CLdeltas, SWcre_LWPdeltas, SWcre_IWPdeltas, SWcre_Tdeltas, SWcre_SHdeltas]
    lw_cre_data = [LWcre_CLdeltas, LWcre_LWPdeltas, LWcre_IWPdeltas, LWcre_Tdeltas, LWcre_SHdeltas]
    width=0.4
    fig, ( ax1, ax2 ) = plt.subplots(2, 1, figsize=(6, 6))
    fig.suptitle('Changes in Cloud Radiative Effect from Ensemble Mean')

    label_added = False
    for i, data in enumerate(sw_cre_data):
        if not label_added:
            for z, j in enumerate(data.items()):
                ax1.scatter(i, j[1], color=colors[z], s=25, label=j[0])
            label_added = True
        else:
            for z, j in enumerate(data.items()):
                ax1.scatter(i, j[1], color=colors[z], s=25)
   
    for i, data in enumerate(lw_cre_data):
        for z, j in enumerate(data.items()):
            ax2.scatter(i, j[1], color=colors[z], s=25)


    ax1.axhline(y=0, label = 'Ensemble Mean', color = 'black', linestyle='--')
    ax2.axhline(y=0, color = 'black', linestyle='--')
    ax1.set_xticks(range(len(labels)))
    ax1.set_xticklabels(labels)
    ax2.set_xticks(range(len(labels)))
    ax2.set_xticklabels(labels)
    ax1.legend(bbox_to_anchor=(1, 1))
    ax1.set_ylabel('$\Delta$SW CRE $Wm^{-2}$')
    ax2.set_ylabel('$\Delta$LW CRE $Wm^{-2}$')
    plt.savefig( location + 'Images/RRTGM/' + label + "_enemble_cre_bias.pdf", format="pdf", bbox_inches='tight')
    plt.show()



    #  CS Plots

    sw_cs_data = [SWcs_CLdeltas, SWcs_LWPdeltas, SWcs_IWPdeltas, SWcs_Tdeltas, SWcs_SHdeltas]
    lw_cs_data = [LWcs_CLdeltas, LWcs_LWPdeltas, LWcs_IWPdeltas, LWcs_Tdeltas, LWcs_SHdeltas]
    width=0.4
    fig, ( ax1, ax2 ) = plt.subplots(2, 1, figsize=(6, 6))
    fig.suptitle('Changes in Outgoing Clear Sky Radiation from Ensemble Mean')

    label_added = False
    for i, data in enumerate(sw_cs_data):
        if not label_added:
            for z, j in enumerate(data.items()):
                ax1.scatter(i, j[1], color=colors[z], s=25, label=j[0])
            label_added = True
        else:
            for z, j in enumerate(data.items()):
                ax1.scatter(i, j[1], color=colors[z], s=25)
   
    for i, data in enumerate(lw_cs_data):
        for z, j in enumerate(data.items()):
            ax2.scatter(i, j[1], color=colors[z], s=25)


    ax1.axhline(y=0, label = 'Ensemble Mean', color = 'black', linestyle='--')
    ax2.axhline(y=0, color = 'black', linestyle='--')
    ax1.set_xticks(range(len(labels)))
    ax1.set_xticklabels(labels)
    ax2.set_xticks(range(len(labels)))
    ax2.set_xticklabels(labels)
    ax1.legend(bbox_to_anchor=(1, 1))
    ax1.set_ylabel('$\Delta$Outgoing SW Clear Sky $Wm^{-2}$')
    ax2.set_ylabel('$\Delta$Outgoing LW Clear Sky $Wm^{-2}$')
    plt.savefig( location + 'Images/RRTGM/' + label + "_enemble_cs_bias.pdf", format="pdf", bbox_inches='tight')
    plt.show()


    #  Liquid and Ice Particle Plots

    labels = ["$r_{liq} = 5 \mu m$", "$r_{liq} = 25 \mu m$", "$r_{liq} = 60 \mu m$", "$r_{ice} = 30 \mu m$", "$r_{ice} = 80 \mu m$", "$r_{ice} = 130 \mu m$"]
    sw_r_data = [SWcre_lr5deltas, SWcre_lr25deltas, SWcre_lr60deltas, SWcre_ir30deltas, SWcre_ir80deltas, SWcre_ir130deltas]
    lw_r_data = [LWcre_lr5deltas, LWcre_lr25deltas, LWcre_lr60deltas, LWcre_ir30deltas, LWcre_ir80deltas, LWcre_ir130deltas]
    width=0.4
    fig, ( ax1, ax2 ) = plt.subplots(2, 1, figsize=(6, 6))
    fig.suptitle('Changes in Cloud Radiative Effect from Ensemble Mean with Particle Size Perturbations')

    for i, data in enumerate(sw_r_data):
        ax1.scatter(i, data, s=25)
   
    for i, data in enumerate(lw_r_data):
        ax2.scatter(i, data, s=25)


    ax1.axhline(y=0, label = 'Ensemble Mean', color = 'black', linestyle='--')
    ax2.axhline(y=0, color = 'black', linestyle='--')
    ax1.set_xticks(range(len(labels)))
    ax1.set_xticklabels(labels)
    ax2.set_xticks(range(len(labels)))
    ax2.set_xticklabels(labels)
    ax1.legend(bbox_to_anchor=(1, 1))
    ax1.set_ylabel('$\Delta$SW CRE $Wm^{-2}$')
    ax2.set_ylabel('$\Delta$LW CRE $Wm^{-2}$')
    plt.savefig( location + 'Images/RRTGM/' + label + "_enemble_r_bias.pdf", format="pdf", bbox_inches='tight')
    plt.show()
