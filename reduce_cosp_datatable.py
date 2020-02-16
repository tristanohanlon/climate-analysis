# -*- coding: utf-8 -*-
"""
Created on Thu Mar 28 18:40:37 2019

@author: Tristan O'Hanlon - University of Auckland & Jonathan Rogers

These functions will create a standard model dataset when parsed:
    
reduce_dataset( directory, filename, save_location, start, end, cl_is_fractional=False )

example:
reduce_dataset( 'gfdl_am4', '_Amon_GFDL-AM4_amip_r1i1p1f1_gr1_198001-201412.nc', 'E:/University/University/MSc/Models/climate-analysis/GFDL-AM4', datetime.datetime( 2000, 1, 1 ), datetime.datetime( 2006, 1, 1 ) )

"""
import numpy as np
import os
from netCDF4 import Dataset
from netCDF4 import date2index
import h5py
import math
from scipy import interpolate
from scipy.interpolate import RectBivariateSpline
from scipy import stats
import datetime
from pprint import pprint
from sklearn.impute import SimpleImputer
import constants
from scipy import ndimage as nd
import matplotlib.pyplot as plt
import openpyxl



def reduce_cosp_datatable( directory, save_location, start, end, rownum, location ):
    book = openpyxl.load_workbook('E:/University/University/MSc/Models/climate-analysis/reduced_data/cloud_data.xlsx')
    sheet = book.active
    # def reduce_dataset( directory, filename, save_location, start, end):
    os.chdir( location + 'Data/' + directory + '/COSP/' )
  
    ################################################---get regional data (time, lat, lon)---################################################

    with Dataset( constants.variable_to_filename( 'cltcalipso' ), 'r') as f:
        raw_lat = constants.extract_data( 'lat', f )
        raw_lon = constants.extract_data( 'lon', f )

        # get indexes for southern ocean -70 to -50 lat

        start_idx = np.abs(raw_lat - (-69.5)).argmin()
        end_idx = np.abs(raw_lat - (-50)).argmin()

        raw_clt_lat_lon = np.mean( constants.extract_data_over_time('cltcalipso', f, start, end ), axis = 0 ) / 100 # average over time


    with Dataset( constants.variable_to_filename( 'cllcalipso' ), 'r') as f:
        raw_clt_l_lat_lon = np.mean( constants.extract_data_over_time('cllcalipso', f, start, end ), axis = 0 ) / 100 # average over time
        raw_clt_l_lat_lon[ raw_clt_l_lat_lon > 1 ] = np.nan

    imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    imp.fit(np.transpose(raw_clt_l_lat_lon))  
    a = imp.transform(np.transpose(raw_clt_l_lat_lon))
    raw_clt_l_lat_lon = np.transpose(a)


    interpolated = interpolate.interp2d(raw_lon, raw_lat, raw_clt_lat_lon, kind = 'cubic')
    clt_lat_lon = interpolated(constants.lon, constants.lat)

    if 'CAM' in directory or 'GISS-E2R' in directory:
        interpolated = interpolate.interp2d(raw_lon, raw_lat[1:], raw_clt_l_lat_lon, kind = 'linear')
        clt_l_lat_lon = interpolated(constants.lon, constants.lat)
    else:
        interpolated = interpolate.interp2d(raw_lon, raw_lat, raw_clt_l_lat_lon, kind = 'linear')
        clt_l_lat_lon = interpolated(constants.lon, constants.lat)

    clt_lat_lon = clt_lat_lon * 100
    clt_l_lat_lon = clt_l_lat_lon * 100

    ######################---get profile data (time, alt40, lat, lon)---###################

    with Dataset( constants.variable_to_filename( 'clcalipso' ), 'r') as f:

        if 'CM6A' in directory:
            raw_alt = constants.extract_data( 'height', f ) / 1000 # in km
        else:
            raw_alt = constants.extract_data( 'alt40', f ) / 1000 # in km



    #-------------create liquid and ice cloud fraction datasets---------------#

    if 'CAM6' in directory or 'CM4' in directory or 'CM6A' in directory:
        with Dataset( constants.variable_to_filename( 'clcalipsoliq' ), 'r') as f:
            clw = np.nanmean( constants.extract_data_over_time( 'clcalipsoliq', f, start, end ), axis = 0 ) / 100 # average over time
        clw[ clw > 1 ] = np.nan

        # clw = clw[5:]
        clw_alt_lat = np.nanmean( clw, axis = -1 ) 

        with Dataset( constants.variable_to_filename( 'clcalipsoice' ), 'r') as f:
            cli = np.nanmean( constants.extract_data_over_time( 'clcalipsoice', f, start, end ), axis = 0 ) / 100 # average over time
        cli[ cli > 1 ] = np.nan
        # cli = cli[5:]
 
        cli_alt_lat = np.nanmean( cli, axis = -1 ) 

        if raw_lat.shape[0] == constants.lat.shape[0]:
            pass
        else:
            if 'CM6A' in directory:
                interpolated = interpolate.interp2d( raw_lat[5:], raw_alt, clw_alt_lat[:,5:], kind = 'cubic')
                clw_alt_lat = interpolated( constants.lat, raw_alt )

                interpolated = interpolate.interp2d( raw_lat[5:], raw_alt, cli_alt_lat[:,5:], kind = 'cubic')
                cli_alt_lat = interpolated( constants.lat, raw_alt )
            else:
                interpolated = interpolate.interp2d( raw_lat, raw_alt, clw_alt_lat, kind = 'cubic')
                clw_alt_lat = interpolated( constants.lat, raw_alt )

                interpolated = interpolate.interp2d( raw_lat, raw_alt, cli_alt_lat, kind = 'cubic')
                cli_alt_lat = interpolated( constants.lat, raw_alt )

        imp = SimpleImputer(missing_values=np.nan, strategy='mean')
        imp.fit(np.transpose(clw_alt_lat))  
        a = imp.transform(np.transpose(clw_alt_lat))
        clw_alt_lat = np.transpose(a)    

        imp = SimpleImputer(missing_values=np.nan, strategy='mean')
        imp.fit(np.transpose(cli_alt_lat))  
        a = imp.transform(np.transpose(cli_alt_lat))
        cli_alt_lat = np.transpose(a)    

        clw[ clw < 1 ] = np.nan
        cli[ cli < 1 ] = np.nan

    # store model name
    sheet.cell(row=rownum, column=1).value = directory
    # store clt
    sheet.cell(row=rownum, column=2).value = constants.global2DMean(clt_lat_lon, constants.lat)
    # store clt SO
    sheet.cell(row=rownum, column=3).value = constants.global2DMean(clt_lat_lon[constants.so_idx_1:constants.so_idx_2], constants.lat[constants.so_idx_1:constants.so_idx_2])
    # store clt_l
    sheet.cell(row=rownum, column=4).value = constants.global2DMean(clt_l_lat_lon, constants.lat)
    # store clt_l SO
    sheet.cell(row=rownum, column=5).value = constants.global2DMean(clt_l_lat_lon[constants.so_idx_1:constants.so_idx_2], constants.lat[constants.so_idx_1:constants.so_idx_2])

    if 'CAM6' in directory or 'CM4' in directory or 'CM6A' in directory:
        # store clw_frac
        sheet.cell(row=rownum, column=6).value = constants.globalalt_latMeanVal(clw_alt_lat, constants.lat)
        # store clw_frac SO
        sheet.cell(row=rownum, column=7).value = constants.globalalt_latMeanVal(clw_alt_lat[:,constants.so_idx_1:constants.so_idx_2], constants.lat[constants.so_idx_1:constants.so_idx_2])
        # store cli_frac
        sheet.cell(row=rownum, column=8).value = constants.globalalt_latMeanVal(cli_alt_lat, constants.lat)
        # store cli_frac SO
        sheet.cell(row=rownum, column=9).value = constants.globalalt_latMeanVal(cli_alt_lat[:,constants.so_idx_1:constants.so_idx_2], constants.lat[constants.so_idx_1:constants.so_idx_2])
    
    book.save('E:/University/University/MSc/Models/climate-analysis/reduced_data/cloud_data.xlsx')

