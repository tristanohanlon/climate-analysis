

import numpy as np
import os
from netCDF4 import Dataset
from netCDF4 import date2index
import h5py
import math
from scipy import interpolate
from scipy.interpolate import RectBivariateSpline
from scipy import stats
import datetime
from pprint import pprint
from sklearn.impute import SimpleImputer
import constants
from scipy import ndimage as nd
import matplotlib.pyplot as plt
import climlab
from matplotlib import rc
rc('text', usetex=True)
from matplotlib.font_manager import FontProperties
fontP = FontProperties()
fontP.set_size('small')
import openpyxl

def radiation(start, end,  location, model, label, min_lat, max_lat, use_aerosol_files, use_surface_albedo, plot_diagnostic_data, save_outputs, use_integrator, liquid_r, ice_r ):
    book = openpyxl.load_workbook( location + 'climate-analysis/reduced_data/radiation_data.xlsx' )
    sheet = book.active

    os.chdir( location + 'Data/' + model )

    #----------------------------- Aerosols ---------------------------#

    # Get aerosol values from model files or set default values if they are
    # not present

    o2=0.21
    ccl4=0.
    cfc22=4.8743326488706363e-11

    # Mole Fraction of Ozone (time, lev, lat, lon)

    variable = 'o3'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        o3 = 0.
    else:
        with Dataset( check_file, 'r') as f:
            o3 = constants.extract_data_over_time( variable, f, start, end )
            o3[o3 > 1] = np.nan
            plev_o3 = constants.extract_data( 'plev', f ) # in Pa


    # Global Mean Mole Fraction of CH4 (time)

    variable = 'ch4global'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        ch4 = 1.557889297535934e-6
    else:
        with Dataset( check_file, 'r') as f:
            ch4 = constants.extract_data_over_time( variable, f, start, end )
            ch4 = np.nanmean( ch4, axis = 0 ) # Average over time


    # Global Mean Mole Fraction of N2O (time)

    variable = 'n2oglobal'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        n2o = 3.007494866529774e-7
    else:
        with Dataset( check_file, 'r') as f:
            n2o = constants.extract_data_over_time( variable, f, start, end )
            n2o = np.nanmean( n2o, axis = 0 ) # Average over time


    # Global Mean Mole Fraction of CFC11 (time)

    variable = 'n2oglobal'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        cfc11 = 1.680e-10
    else:
        with Dataset( check_file, 'r') as f:
            cfc11 = constants.extract_data_over_time( variable, f, start, end )
            cfc11 = np.nanmean( cfc11, axis = 0 ) # Average over time


    # Global Mean Mole Fraction of CFC12 (time)

    variable = 'cfc12global'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        cfc12 = 2.850e-10
    else:
        with Dataset( check_file, 'r') as f:
            cfc12 = constants.extract_data_over_time( variable, f, start, end )
            cfc12 = np.nanmean( cfc12, axis = 0 ) # Average over time


    # Global Mean Mole Fraction of CFC113 (time)

    variable = 'cfc113global'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        cfc113 = 1.737268993839836e-11
    else:
        with Dataset( check_file, 'r') as f:
            cfc113 = constants.extract_data_over_time( variable, f, start, end )
            cfc113 = np.nanmean( cfc113, axis = 0 ) # Average over time


    # Total Atmospheric Mass (kg) of CO2 

    variable = 'co2mass'
    check_file = constants.variable_to_filename( variable )
    if check_file == None:
        co2mass = 400.e-6
    else:
        with Dataset( check_file, 'r') as f:
            co2mass = constants.extract_data_over_time( variable, f, start, end )
            co2mass = np.nanmean( co2mass, axis = 0 ) # Average over time


    #----------------------------- Cloud variables -----------------------------#

    # Get layer cloud fraction and pressure level variables - (time, level, lat, lon)

    with Dataset( constants.variable_to_filename( 'cl' ), 'r') as f:
        lat = constants.extract_data( 'lat', f )
        # Define latitude confinement parsed from function
        lat_confine_1 = np.abs(lat - (min_lat)).argmin()
        lat_confine_2 = np.abs(lat - (max_lat)).argmin()
        lat = lat[lat_confine_1:lat_confine_2]

        lon = constants.extract_data( 'lon', f )
        cl = constants.extract_data_over_time( 'cl', f, start, end ) / 100
        if 'CM4' in model or 'IPSL' in model:
            a = constants.extract_data( 'ap', f )
            b = constants.extract_data( 'b', f )
        else:
            a = constants.extract_data( 'a', f )
            b = constants.extract_data( 'b', f )
            p0 = np.array(f.variables['p0'][:])
            a = a*p0
        cl = cl[:,:,lat_confine_1:lat_confine_2,:]


    # Get surface pressure - (time, lat, lon)

    with Dataset( constants.variable_to_filename( 'ps' ), 'r') as f:
        ps = constants.extract_data_over_time( 'ps', f, start, end )
        ps = np.nanmean( ps, axis = 0 ) # average over time


    # Get surface temperature - (time, lat, lon)

    with Dataset( constants.variable_to_filename( 'ts' ), 'r') as f:
        ts = constants.extract_data_over_time( 'ts', f, start, end )
        ts = ts[:,lat_confine_1:lat_confine_2,:]


    # Get specific humidity - (time, lev, lat, lon)

    with Dataset( constants.variable_to_filename( 'hus' ), 'r') as f:
        hus = constants.extract_data_over_time( 'hus', f, start, end )
        hus[hus > 1] = np.nan
        plev_h = constants.extract_data( 'plev', f ) # in Pa
        hus = hus[:,:,lat_confine_1:lat_confine_2,:]


    # Convert pressure level variables to pressure

    a = np.swapaxes(np.swapaxes(np.tile(a, (np.shape(ps)[0], np.shape(ps)[1], 1) ), 0, 2), 1, 2)
    b = np.swapaxes(np.swapaxes(np.tile(b, (np.shape(ps)[0], np.shape(ps)[1], 1) ), 0, 2), 1, 2)
    ps = np.tile(ps, (np.shape(b)[0], 1, 1) )
    p = (a + b*ps) # in Pa
    p = p[:,lat_confine_1:lat_confine_2,:]


    # Get Cloud liquid water mass fraction in air (kg/kg) - (time, level, lat, lon)

    with Dataset( constants.variable_to_filename( 'clw' ), 'r') as f:
        clw = constants.extract_data_over_time( 'clw', f, start, end )
        clw = clw[:,:,lat_confine_1:lat_confine_2,:]


    # Get Cloud ice water mass fraction in air (kg/kg) - (time, level, lat, lon)

    with Dataset( constants.variable_to_filename( 'cli' ), 'r') as f:
        cli = constants.extract_data_over_time( 'cli', f, start, end )
        cli = cli[:,:,lat_confine_1:lat_confine_2,:]


    # Get temperature in atmosphere levels (time, level, lat, lon)

    with Dataset( constants.variable_to_filename( 'ta' ), 'r' ) as f:
        ta = constants.extract_data_over_time( 'ta', f, start, end )
        ta[ta > 500] = np.nan
        plev_t = constants.extract_data( 'plev', f ) # in Pa
        ta = ta[:,:,lat_confine_1:lat_confine_2,:]

    # Confine O3 data to set latitudes

        o3 = o3[:,:,lat_confine_1:lat_confine_2,:]



    #-------------------------- Get top of atmosphere radiative flux data ------------------------#

    # Incoming SW at TOA

    with Dataset( constants.variable_to_filename( 'rsdt' ), 'r') as f:
        rsdt = constants.extract_data('rsdt', f )
        rsdt = rsdt[:,lat_confine_1:lat_confine_2,:]


    # Outgoing SW at TOA

    with Dataset( constants.variable_to_filename( 'rsut' ), 'r') as f:
        rsut = constants.extract_data('rsut', f )
        rsut = rsut[:,lat_confine_1:lat_confine_2,:]


    # Outgoing SW at TOA assuming clear sky

    with Dataset( constants.variable_to_filename( 'rsutcs' ), 'r') as f:
        rsutcs = constants.extract_data('rsutcs', f )
        rsutcs = rsutcs[:,lat_confine_1:lat_confine_2,:]


    # Outgoing LW at TOA

    with Dataset( constants.variable_to_filename( 'rlut' ), 'r') as f:
        rlut = constants.extract_data('rlut', f )
        rlut = rlut[:,lat_confine_1:lat_confine_2,:]


    # Outgoing LW at TOA assuming clear sky

    with Dataset( constants.variable_to_filename( 'rlutcs' ), 'r') as f:
        rlutcs = constants.extract_data('rlutcs', f )
        rlutcs = rlutcs[:,lat_confine_1:lat_confine_2,:]


    # Net TOA flux

    with Dataset( constants.variable_to_filename( 'rtmt' ), 'r') as f:
        rtmt = constants.extract_data('rtmt', f )
        rtmt = rtmt[:,lat_confine_1:lat_confine_2,:]


    #-------------------------- Surface radiative flux data --------------------------#

    # Incoming SW at surface

    with Dataset( constants.variable_to_filename( 'rsds' ), 'r') as f:
        rsds = constants.extract_data('rsds', f )
        rsds = rsds[:,lat_confine_1:lat_confine_2,:]


    # Incoming SW at surface assuming clear sky

    with Dataset( constants.variable_to_filename( 'rsdscs' ), 'r') as f:
        rsdscs = constants.extract_data('rsdscs', f )
        rsdscs = rsdscs[:,lat_confine_1:lat_confine_2,:]


    # Outgoing SW at surface

    with Dataset( constants.variable_to_filename( 'rsus' ), 'r') as f:
        rsus = constants.extract_data('rsus', f )
        rsus = rsus[:,lat_confine_1:lat_confine_2,:]


    # Outgoing SW at surface assuming clear sky

    with Dataset( constants.variable_to_filename( 'rsuscs' ), 'r') as f:
        rsuscs = constants.extract_data('rsuscs', f )
        rsuscs = rsuscs[:,lat_confine_1:lat_confine_2,:]


    # Outgoing LW at surface

    with Dataset( constants.variable_to_filename( 'rlus' ), 'r') as f:
        rlus = constants.extract_data('rlus', f )
        rlus = rlus[:,lat_confine_1:lat_confine_2,:]


    # Incoming LW at surface

    with Dataset( constants.variable_to_filename( 'rlds' ), 'r') as f:
        rlds = constants.extract_data('rlds', f )
        rlds = rlds[:,lat_confine_1:lat_confine_2,:]


    # Incoming LW at surface assuming clear sky

    with Dataset( constants.variable_to_filename( 'rldscs' ), 'r') as f:
        rldscs = constants.extract_data('rldscs', f )
        rldscs = rldscs[:,lat_confine_1:lat_confine_2,:]


    # Determine albedo values at surface and TOA

    albedo_toa = rsut / rsdt
    albedo_toa_cs = rsutcs / rsdt
    albedo_surface = rsus / rsds
    albedo_surface_cs = rsuscs / rsdscs


    #-------------------------- Prepare variables for time integration or average the variables over time --------------------------#

    # Model radiation data for comparison and diagnostics

    rsdt = np.nanmean( rsdt, axis = 0 )
    rsut = np.nanmean( rsut, axis = 0 )
    rsutcs = np.nanmean( rsutcs, axis = 0 )
    rlut = np.nanmean( rlut, axis = 0 )
    rlutcs = np.nanmean( rlutcs, axis = 0 )
    rtmt = np.nanmean( rtmt, axis = 0 )

    rsds = np.nanmean( rsds, axis = 0 )
    rsdscs = np.nanmean( rsdscs, axis = 0 )
    rsus = np.nanmean( rsus, axis = 0 )
    rsuscs = np.nanmean( rsuscs, axis = 0 )
    rlus = np.nanmean( rlus, axis = 0 )
    rlds = np.nanmean( rlds, axis = 0 )
    rldscs = np.nanmean( rldscs, axis = 0 )

    albedo_toa = np.nanmean( albedo_toa, axis = 0 )
    albedo_toa_cs = np.nanmean( albedo_toa_cs, axis = 0 )
    albedo_surface = np.nanmean( albedo_surface, axis = 0 )
    albedo_surface_cs = np.nanmean( albedo_surface_cs, axis = 0 )

    # Get number of months to integrate over

    dt = np.size( cl, 0 )

    if use_integrator == False:
        cl = np.nanmean( cl, axis = 0 )
        clw = np.nanmean( clw, axis = 0 )
        cli = np.nanmean( cli, axis = 0 )
        ta = np.nanmean( ta, axis = 0 )
        ts = np.nanmean( ts, axis = 0 )
        hus = np.nanmean( hus, axis = 0 )
        o3 = np.nanmean( o3, axis = 0 )

    #-------------------------- Average variables over longitude --------------------------#

    cl = np.nanmean( cl, axis = -1 )
    clw = np.nanmean( clw, axis = -1 )
    cli = np.nanmean( cli, axis = -1 )
    ta = np.nanmean( ta, axis = -1 )
    ts = np.nanmean( ts, axis = -1 )
    hus = np.nanmean( hus, axis = -1 )
    plev = np.nanmean( p, axis = -1 )

    rsdt = np.nanmean( rsdt, axis = -1 )
    rsut = np.nanmean( rsut, axis = -1 )
    rsutcs = np.nanmean( rsutcs, axis = -1 )
    rlut = np.nanmean( rlut, axis = -1 )
    rlutcs = np.nanmean( rlutcs, axis = -1 )
    rtmt = np.nanmean( rtmt, axis = -1 )

    rsds = np.nanmean( rsds, axis = -1 )
    rsdscs = np.nanmean( rsdscs, axis = -1 )
    rsus = np.nanmean( rsus, axis = -1 )
    rsuscs = np.nanmean( rsuscs, axis = -1 )
    rlus = np.nanmean( rlus, axis = -1 )
    rlds = np.nanmean( rlds, axis = -1 )
    rldscs = np.nanmean( rldscs, axis = -1 )

    albedo_toa = np.nanmean( albedo_toa, axis = -1 )
    albedo_toa_cs = np.nanmean( albedo_toa_cs, axis = -1 )
    albedo_surface = np.nanmean( albedo_surface, axis = -1 )
    albedo_surface_cs = np.nanmean( albedo_surface_cs, axis = -1 )

    o3 = np.nanmean( o3, axis = -1 )


    # Get layer pressure levels (Pa)
    p = constants.global3DMean(p, lat)

    #-------------------------- Interpolate variables --------------------------#

    # Interpolate temperature

    imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    imp.fit(np.transpose(ta))  
    ta = imp.transform(np.transpose(ta))
    ta = np.transpose(ta)   
    if plev_t.shape[0] > ta.shape[0]:
        plev_t = plev_t[plev_t.shape[0] - ta.shape[0]:] # reshape alt_temp if not equal
    interpolated = interpolate.interp2d( lat, plev_t, ta, kind = 'linear')
    ta = np.flip( interpolated( lat, p ), axis = 0 )

    # Interpolate humidity

    imp.fit(np.transpose(hus))  
    hus = imp.transform(np.transpose(hus))
    hus = np.transpose(hus)   
    if plev_h.shape[0] > hus.shape[0]:
        plev_h = plev_h[plev_h.shape[0] - hus.shape[0]:] # reshape alt_temp if not equal
    interpolated = interpolate.interp2d( lat, plev_h, hus, kind = 'linear')
    hus = np.flip( interpolated( lat,p ), axis = 0 )

    # Interpolate O3

    imp.fit(np.transpose(o3))  
    o3 = imp.transform(np.transpose(o3))
    o3 = np.transpose(o3)   
    if plev_o3.shape[0] > o3.shape[0]:
        plev_o3 = plev_o3[plev_o3.shape[0] - o3.shape[0]:] # reshape alt_temp if not equal
    interpolated = interpolate.interp2d( lat, plev_o3, o3, kind = 'linear')
    o3 = np.flip( interpolated( lat,p ), axis = 0 )


    #-------------------------- Convert mixing ratios to water paths, set particle sizes --------------------------#

    # Convert mixing ratio (kg/kg) into cloud water content in (g/m3)

    clwc = constants.mix_ratio_to_water_content( clw, ta, plev )
    ciwc = constants.mix_ratio_to_water_content( cli, ta, plev )
    alt = constants.p_to_alt( p ) # in km

    # Convert cloud water content to water path (g/m2)
    # WP = WC * deltaz
    clwp = constants.wc_to_wp( clwc, alt )
    ciwp = constants.wc_to_wp( ciwc, alt )

    # Set liquid and ice droplet and particle diameters

    r_liq = np.zeros((p.shape[0], lat.shape[0]))
    r_ice = np.zeros((p.shape[0], lat.shape[0]))

    r_liq[:] = liquid_r
    r_ice[:] = ice_r


    #-------------------------- diagnostic plots --------------------------#

    # Check for anomolies in the model input data - save as a pdf

    if plot_diagnostic_data == True:

        fig, axs = plt.subplots(4, 2, figsize=(11, 16))
        fig.suptitle('Diagnostics - ' + model)
        xx, yy = np.meshgrid(lat, alt)
    
        axs[0, 0].set_title('O3 Mole Fraction')
        axs[0, 0].set_ylabel('Altitude (km)')
        f = axs[0, 0].contourf(xx, yy, o3, 5, cmap='coolwarm', alpha=0.4)
        plt.colorbar(f, ax=axs[0, 0], label='O3 Mole Fraction')

        axs[0, 1].set_title('Atmosphere Temperature $K$')
        f = axs[0, 1].contourf(xx, yy, ta, 5, cmap='coolwarm', alpha=0.4)
        plt.colorbar(f, ax=axs[0, 1], label='Atmosphere Temperature $K$')

        axs[1, 0].set_title('Atmosphere Specific Humidity $kg/kg$')
        axs[1, 0].set_ylabel('Altitude (km)')
        f = axs[1, 0].contourf(xx, yy, hus, 5, cmap='coolwarm', alpha=0.4)
        plt.colorbar(f, ax=axs[1, 0], label='Atmosphere Specific Humidity $kg/kg$')
        
        axs[1, 1].set_title('Cloud Fraction')
        f = axs[1, 1].contourf(xx, yy, cl, 5, cmap='coolwarm', alpha=0.4)
        plt.colorbar(f, ax=axs[1, 1], label='Cloud Fraction')

        axs[2, 0].set_title('Cloud Liquid Water Path')
        axs[2, 0].set_ylabel('Altitude (km)')
        f = axs[2, 0].contourf(xx, yy, clwp, 5, cmap='coolwarm', alpha=0.4)
        plt.colorbar(f, ax=axs[2, 0], label='Cloud Liquid Water Path $gm^{-2}$')
        
        axs[2, 1].set_title('Cloud Ice Water Path')
        f = axs[2, 1].contourf(xx, yy, ciwp, 5, cmap='coolwarm', alpha=0.4)
        plt.colorbar(f, ax=axs[2, 1], label='Cloud Ice Water Path $gm^{-2}$')

        axs[3, 0].plot( lat, ( rsdt - rsut ), label='SW Net', color = 'black', linestyle='-')
        axs[3, 0].plot( lat, ( rsutcs - rsut ), label='SW - Clouds', color = 'blue', linestyle='-' )
        axs[3, 0].plot( lat, ( rsdt - rsutcs ), label='SW - Clear Sky', color = 'black', linestyle=':')
        axs[3, 0].set_ylabel('Radiation $Wm^{-2}$')
        axs[3, 0].set_xlabel('Latitude')
        axs[3, 0].set_title ('Model Output - Absorbed Shortwave Flux (ASR)')
        axs[3, 0].legend(loc='best', prop=fontP);

        axs[3, 1].plot( lat, rlut, label='LW Net', color = 'black', linestyle='-')
        axs[3, 1].plot( lat, ( rlut - rlutcs ), label='LW - Clouds', color = 'blue', linestyle='-')
        axs[3, 1].plot( lat, rlutcs, label='LW - Clear Sky', color = 'black', linestyle=':' )
        axs[3, 1].set_ylabel('Radiation $Wm^{-2}$')
        axs[3, 1].set_xlabel('Latitude')
        axs[3, 1].set_title ('Model Output - Outgoing Longwave Flux (OLR)')
        axs[3, 1].legend(loc='best', prop=fontP);

        # axs[3, 0].plot( lat, rsds, label='SW Down', color = 'black', linestyle='-')
        # axs[3, 0].plot( lat, rsdscs, label='SW Down - Clear Sky', color = 'black', linestyle='--' )
        # axs[3, 0].plot( lat, rsus, label='SW Up', color = 'blue', linestyle='-')
        # axs[3, 0].plot( lat, rsuscs, label='SW Up - Clear Sky', color = 'blue', linestyle='--' )
        # axs[3, 0].plot( lat, rlus, label='LW Up', color = 'green', linestyle='-')
        # axs[3, 0].plot( lat, rlds, label='LW Down', color = 'red', linestyle='-' )
        # axs[3, 0].plot( lat, rldscs, label='LW Down - Clear Sky', color = 'red', linestyle='--' )
        # axs[3, 0].set_ylabel('Radiation $Wm^{-2}$')
        # axs[3, 0].set_xlabel('Latitude')
        # axs[3, 0].set_title ('Model Output - Absorbed Shortwave Flux (ASR)')
        # axs[3, 0].legend(loc='best', prop=fontP);

        # axs[3, 1].plot( lat, rsdt, label='SW Down', color = 'black', linestyle='-')
        # axs[3, 1].plot( lat, rsut, label='SW Up', color = 'blue', linestyle='-')
        # axs[3, 1].plot( lat, rsutcs, label='SW Up - Clear Sky', color = 'blue', linestyle='--' )
        # axs[3, 1].plot( lat, rlut, label='LW Up', color = 'red', linestyle='-')
        # axs[3, 1].plot( lat, rlutcs, label='LW Up - Clear Sky', color = 'red', linestyle='--' )
        # axs[3, 1].set_ylabel('Radiation $Wm^{-2}$')
        # axs[3, 1].set_xlabel('Latitude')
        # axs[3, 1].set_title ('Model Output - Outgoing Longwave Flux (OLR)')
        # axs[3, 1].legend(loc='best', prop=fontP);

        plt.savefig( location + '/Images/RRTGM/diagnostics/' + "diagnostics_" + label + ".pdf", format="pdf")


    #-------------------------- Input variables into radiative transfer code --------------------------#

    sfc, atm = climlab.domain.zonal_mean_column(lat=lat, lev=p)

    # Surface variables
    insolation_field = climlab.domain.Field(rsdt, domain=sfc)
    albedo_surface_field = climlab.domain.Field(albedo_surface, domain=sfc)  
    albedo_surface_cs_field = climlab.domain.Field(albedo_surface_cs, domain=sfc)  
    albedo_toa_field = climlab.domain.Field(albedo_toa, domain=sfc)  
    albedo_toa_cs_field = climlab.domain.Field(albedo_toa_cs, domain=sfc)  
    ts_field = climlab.domain.Field(ts, domain=sfc)

    # Atmosphere layer variables
    ta_field = climlab.domain.Field(np.transpose(ta), domain=atm)
    specific_humidity_field = climlab.domain.Field(np.transpose(hus), domain=atm) # kg/kg
    cldfrac_field = climlab.domain.Field(np.transpose(cl), domain=atm) # fraction
    o3_field = climlab.domain.Field(np.transpose(o3*1000), domain=atm)
    r_liq_field = climlab.domain.Field(np.transpose(r_liq), domain=atm) # Cloud water drop effective radius (microns)
    r_ice_field = climlab.domain.Field(np.transpose(r_ice), domain=atm) # Cloud ice particle effective size (microns)        

    clw_field = climlab.domain.Field(np.transpose(clw), domain=atm) # mixing ratio kg/kg
    ciw_field = climlab.domain.Field(np.transpose(cli), domain=atm) # mixing ratio kg/kg

    clwc_field = climlab.domain.Field(np.transpose(clwc), domain=atm) # content g/m3
    ciwc_field = climlab.domain.Field(np.transpose(ciwc), domain=atm) # content g/m3

    clwp_field = climlab.domain.Field(np.transpose(clwp), domain=atm) # water path g/m2
    ciwp_field = climlab.domain.Field(np.transpose(ciwp), domain=atm) # water path g/m2


    # Dictionary of volumetric mixing ratios. Default values supplied if None
    # if absorber_vmr = None then ozone will be interpolated to the model grid from a climatology file, or set to zero if ozone_file = None.
    if use_aerosol_files == True:
        absorber = {'O3': o3_field, 'CO2': 400.e-6, 'CH4':ch4, 'N2O':n2o, 'O2': o2,'CCL4':ccl4, 
            'CFC11':cfc11, 'CFC12':cfc12, 'CFC113':cfc113, 'CFC22':cfc22}
    else:
        absorber = {'O3': o3_field, 'CO2': 400.e-6, 'CH4':1.557889297535934e-6, 'N2O':3.007494866529774e-7, 'O2': o2,'CCL4':ccl4, 
            'CFC11':1.680e-10, 'CFC12':2.850e-10, 'CFC113':1.737268993839836e-11, 'CFC22':cfc22}

    # check if the model should use the given surface albedo from file or not
    if use_surface_albedo == False:
        albedo_surface_field = None
    

    #  state variables (Air and surface temperature)
    state = {'Tatm': ta_field, 'Ts': ts_field}

    #-------------------------- Execute Control Data --------------------------#

    sw_down = np.zeros_like(ta_field)
    sw_down_clr = np.zeros_like(ta_field)
    sw_up = np.zeros_like(ta_field)
    sw_up_clr = np.zeros_like(ta_field)
    sw_net = np.zeros_like(ta_field)

    lw_down = np.zeros_like(ta_field)
    lw_down_clr = np.zeros_like(ta_field)
    lw_up_clr = np.zeros_like(ta_field)
    lw_up = np.zeros_like(ta_field)
    lw_net = np.zeros_like(ta_field)

    net_ASR = np.zeros_like(ts_field)
    net_ASRcld = np.zeros_like(ts_field)
    net_ASRclr = np.zeros_like(ts_field)  

    net_OLR = np.zeros_like(ts_field)
    net_OLRcld = np.zeros_like(ts_field)
    net_OLRclr = np.zeros_like(ts_field)  

    for i in range(lat.size):
        control_rad = climlab.radiation.RRTMG(name='Radiation', 
                                        state=state, 
                                        specific_humidity=specific_humidity_field, 
                                        absorber_vmr=absorber,
                                        # albedo=albedo_surface_field, 
                                        # insolation = insolation_field, 
                                        cldfrac=cldfrac_field, 
                                        r_liq=r_liq_field, 
                                        r_ice=r_ice_field, 
                                        clwp=clwp_field, 
                                        ciwp=ciwp_field
                                        )

        control_rad.compute_diagnostics()

        # sw_down[i] = control_rad.SW_flux_down
        # sw_down_clr[i] = control_rad.SW_flux_down_clr
        # sw_up[i] = control_rad.SW_flux_up
        # sw_up_clr[i] = control_rad.SW_flux_up_clr
        # sw_net[i] = control_rad.SW_flux_net

        # lw_down[i] = control_rad.LW_flux_down
        # lw_down_clr[i] = control_rad.LW_flux_down_clr
        # lw_up_clr[i] = control_rad.LW_flux_up_clr
        # lw_up[i] = control_rad.LW_flux_up
        # lw_net[i] = control_rad.LW_flux_net

        net_ASR[i] = control_rad.ASR
        net_ASRcld[i] = control_rad.ASRcld
        net_ASRclr[i] = control_rad.ASRclr   

        net_OLR[i] = control_rad.OLR
        net_OLRcld[i] = control_rad.OLRcld
        net_OLRclr[i] = control_rad.OLRclr   


    #-------------------------- Plots --------------------------#
    if save_outputs == True:

        # SW plots
        fig, axs = plt.subplots(3, 2, figsize=(11, 11))
        fig.suptitle('SW Flux - ' + model)
    
        axs[0, 0].set_title('SW Down - Clear Sky')
        axs[0, 0].set_ylabel('Altitude (km)')
        f = axs[0, 0].contourf(lat, alt[:15], np.transpose(sw_down_clr[:,1:16]), 5, cmap='Greens', alpha=0.4)
        plt.colorbar(f, ax=axs[0, 0], label='$Wm^{-2}$')

        axs[0, 1].set_title('SW Down - Clouds')
        f = axs[0, 1].contourf(lat, alt[:15], np.transpose(sw_down[:,1:16]), 5, cmap='Greens', alpha=0.4)
        plt.colorbar(f, ax=axs[0, 1], label='$Wm^{-2}$')

        axs[1, 0].set_title('SW Up - Clear Sky')
        axs[1, 0].set_ylabel('Altitude (km)')
        f = axs[1, 0].contourf(lat, alt[:15], np.transpose(sw_up_clr[:,1:16]), 5, cmap='Reds', alpha=0.4)
        plt.colorbar(f, ax=axs[1, 0], label='$Wm^{-2}$')
        
        axs[1, 1].set_title('SW Up - Clouds')
        f = axs[1, 1].contourf(lat, alt[:20], np.transpose(sw_up[:,1:21]), 5, cmap='Reds', alpha=0.4)
        plt.colorbar(f, ax=axs[1, 1], label='$Wm^{-2}$')

        axs[2, 0].set_title('SW Net Flux $Wm^{-2}$ - ' + model)
        axs[2, 0].set_ylabel('Altitude (km)')
        axs[2, 0].set_xlabel('Latitude')
        f = axs[2, 0].contourf( lat, alt[:20], np.transpose(sw_net[:,1:21]), 5, cmap='Blues' )
        plt.colorbar(f, ax=axs[2, 0], label='SW Net Flux $Wm^{-2}$')
        
        axs[2, 1].plot( lat, net_ASR, label='Net ASR', color = 'black', linestyle='-')
        axs[2, 1].plot( lat, net_ASRcld, label='ASR With Clouds', color = 'blue', linestyle='-' )
        axs[2, 1].plot( lat, net_ASRclr, label='ASR Clear Sky', color = 'black', linestyle=':' )
        axs[2, 1].set_ylabel('Radiation $Wm^{-2}$')
        axs[2, 1].set_xlabel('Latitude')
        axs[2, 1].set_title('Absorbed Solar Radiation - ' + model)
        axs[2, 1].legend(loc='best', prop=fontP);

        plt.savefig( location + '/Images/RRTGM/' + "sw_output_" + label + ".pdf", format="pdf")



        # LW plots
        fig, axs = plt.subplots(3, 2, figsize=(11, 11))
        fig.suptitle('LW Flux - ' + model)
    
        axs[0, 0].set_title('LW Down - Clear Sky')
        axs[0, 0].set_ylabel('Altitude (km)')
        f = axs[0, 0].contourf(lat, alt[:20], np.transpose(lw_down_clr[:,1:21]), 5, cmap='Greens', alpha=0.4)
        plt.colorbar(f, ax=axs[0, 0], label='$Wm^{-2}$')

        axs[0, 1].set_title('LW Down - Clouds')
        f = axs[0, 1].contourf(lat, alt[:20], np.transpose(lw_down[:,1:21]), 5, cmap='Greens', alpha=0.4)
        plt.colorbar(f, ax=axs[0, 1], label='$Wm^{-2}$')

        axs[1, 0].set_title('LW Up - Clear Sky')
        axs[1, 0].set_ylabel('Altitude (km)')
        f = axs[1, 0].contourf(lat, alt[:20], np.transpose(lw_up_clr[:,1:21]), 5, cmap='Reds', alpha=0.4)
        plt.colorbar(f, ax=axs[1, 0], label='$Wm^{-2}$')
        
        axs[1, 1].set_title('LW Up - Clouds')
        f = axs[1, 1].contourf(lat, alt[:20], np.transpose(lw_up[:,1:21]), 5, cmap='Reds', alpha=0.4)
        plt.colorbar(f, ax=axs[1, 1], label='$Wm^{-2}$')

        axs[2, 0].set_title('LW Net Flux $Wm^{-2}$ - ' + model)
        axs[2, 0].set_ylabel('Altitude (km)')
        axs[2, 0].set_xlabel('Latitude')
        f = axs[2, 0].contourf( lat, alt[:20], np.transpose(lw_net[:,1:21]), 5, cmap='Blues' )
        plt.colorbar(f, ax=axs[2, 0], label='LW Net Flux $Wm^{-2}$')
        
        axs[2, 1].plot( lat, net_OLR, label='Net OLR', color = 'black', linestyle='-')
        axs[2, 1].plot( lat, net_OLRcld, label='OLR With Clouds', color = 'blue', linestyle='-' )
        axs[2, 1].plot( lat, net_OLRclr, label='OLR Clear Sky', color = 'black', linestyle=':' )
        axs[2, 1].set_ylabel('Radiation $Wm^{-2}$')
        axs[2, 1].set_xlabel('Latitude')
        axs[2, 1].set_title('Outgoing Longwave Radiation - ' + model)
        axs[2, 1].legend(loc='best', prop=fontP);

        plt.savefig( location + '/Images/RRTGM/' + "lw_output_" + label + ".pdf", format="pdf")


        # Store global means in excel file
        start_row = int(label) + 2
        sheet.cell(row=start_row, column=1).value = label
        sheet.cell(row=start_row, column=2).value = model
        sheet.cell(row=start_row, column=3).value = liquid_r
        sheet.cell(row=start_row, column=4).value = ice_r
        sheet.cell(row=start_row, column=5).value = use_aerosol_files
        sheet.cell(row=start_row, column=6).value = constants.globalMean(np.array(net_ASR)[:,0], lat)
        sheet.cell(row=start_row, column=7).value = constants.globalMean(np.array(net_OLR)[:,0], lat)
        sheet.cell(row=start_row, column=8).value = constants.globalalt_latMeanVal(np.transpose(np.array(sw_net)), lat)
        sheet.cell(row=start_row, column=9).value = constants.globalalt_latMeanVal(np.transpose(np.array(lw_net)), lat)
        sheet.cell(row=start_row, column=10).value = constants.globalMean(np.array(net_ASRclr)[:,0], lat)
        sheet.cell(row=start_row, column=11).value = constants.globalMean(np.array(net_ASRcld)[:,0], lat)
        sheet.cell(row=start_row, column=12).value = constants.globalMean(np.array(net_OLRclr)[:,0], lat)
        sheet.cell(row=start_row, column=13).value = constants.globalMean(np.array(net_OLRcld)[:,0], lat)
        sheet.cell(row=start_row, column=14).value = constants.globalalt_latMeanVal(np.transpose(np.array(sw_down_clr)), lat)
        sheet.cell(row=start_row, column=15).value = constants.globalalt_latMeanVal(np.transpose(np.array(sw_down)), lat)
        sheet.cell(row=start_row, column=16).value = constants.globalalt_latMeanVal(np.transpose(np.array(sw_up_clr)), lat)
        sheet.cell(row=start_row, column=17).value = constants.globalalt_latMeanVal(np.transpose(np.array(sw_up)), lat)
        sheet.cell(row=start_row, column=18).value = constants.globalalt_latMeanVal(np.transpose(np.array(lw_down_clr)), lat)
        sheet.cell(row=start_row, column=19).value = constants.globalalt_latMeanVal(np.transpose(np.array(lw_down)), lat)
        sheet.cell(row=start_row, column=20).value = constants.globalalt_latMeanVal(np.transpose(np.array(lw_up_clr)), lat)
        sheet.cell(row=start_row, column=21).value = constants.globalalt_latMeanVal(np.transpose(np.array(lw_up)), lat)
  
        book.save( location + 'climate-analysis/reduced_data/radiation_data.xlsx' )


