# -*- coding: utf-8 -*-
"""
Created on Thu Mar 28 18:40:37 2019

@author: Tristan O'Hanlon - University of Auckland & Jonathan Rogers

These functions will create a standard model dataset when parsed:
    
reduce_dataset( directory, filename, save_location, start, end, cl_is_fractional=False )

example:
reduce_dataset( 'gfdl_am4', '_Amon_GFDL-AM4_amip_r1i1p1f1_gr1_198001-201412.nc', 'E:/University/University/MSc/Models/climate-analysis/GFDL-AM4', datetime.datetime( 2000, 1, 1 ), datetime.datetime( 2006, 1, 1 ) )

if cl_is_fractional=True - the values get dived by 100
"""
import numpy as np
import os
from netCDF4 import Dataset
from netCDF4 import date2index
import h5py
import math
from scipy import interpolate
from scipy.interpolate import RectBivariateSpline
from scipy import stats
import datetime
from pprint import pprint
from sklearn.impute import SimpleImputer
import constants
from scipy import ndimage as nd
import matplotlib.pyplot as plt
import openpyxl


def reduce_dataset( directory, filename, save_location, start, end, rownum ):

    book = openpyxl.load_workbook('E:/University/University/MSc/Models/climate-analysis/reduced_data/cloud_data.xlsx')
    sheet = book.active
    # def reduce_dataset( directory, filename, save_location, start, end):
    os.chdir( directory )
  
    ################################################---get regional data (time, lat, lon)---################################################


#  or 'RCP45-IPSL' in directory or 'RCP45-MRI' in directory or 'future4K-MIROC6' in directory or 'future4K-MRI' in directory or 'SSP245-MIROC6' in directory
    if 'RCP45-GFDL' in directory:
        with Dataset( 'clt' + filename, 'r') as f:
            raw_lat = constants.extract_data( 'lat', f )
            raw_lon = constants.extract_data( 'lon', f )
            # clt = np.mean( constants.extract_data('clt', f ), axis = 0 ) # average over time

        # get indexes for southern ocean -70 to -50 lat

        start_idx = np.abs(raw_lat - (-69.5)).argmin()
        end_idx = np.abs(raw_lat - (-50)).argmin()


        #---get radiative flux data---#

        with Dataset( 'rsdt' + filename, 'r') as f:
            rsdt = np.mean( constants.extract_data('rsdt', f ), axis = 0 ) # average over time
            
        with Dataset( 'rsut' + filename, 'r') as f:
            rsut = np.mean( constants.extract_data('rsut', f ), axis = 0 ) # average over time           

        with Dataset( 'rsutcs' + filename, 'r') as f:
            rsutcs = np.mean( constants.extract_data('rsutcs', f ), axis = 0 ) # average over time           
        
        with Dataset( 'rlut' + filename, 'r') as f:
            rlut = np.mean( constants.extract_data('rlut', f ), axis = 0 ) # average over time           

        with Dataset( 'rlutcs' + filename, 'r') as f:
            rlutcs = np.mean( constants.extract_data('rlutcs', f ), axis = 0 ) # average over time           
        
        if 'MRI' in directory:
            rtmt = ( rsdt ) - ( rsut + rlut ) # https://www4.uwsp.edu/geo/faculty/lemke/geog101/lectures/02_radiation_energy_balance.html
        else:
            with Dataset( 'rtmt' + filename, 'r') as f:
                rtmt = np.mean( constants.extract_data('rtmt', f ), axis = 0 ) # average over time
        
                
        #---create albedo data---#
        
        albedo_reg = rsut / rsdt
        cre_sw_reg = rsutcs - rsut
        cre_lw_reg = rlutcs - rlut
        
        # with Dataset( 'cl' + filename, 'r') as f:
        #     cl = constants.extract_data( 'cl',f )
        #     cl = np.mean( cl, axis = 0 ) # average over time
        #     if directory == 'CMIP5-AMIP-CESM1-CAM5':
        #         pass
        #     else:
        #         cl = cl / 100
        #     if 'GFDL-AM4' in directory or 'IPSL' in directory or 'GFDL-CM4' in directory:
        #         a = constants.extract_data('ap', f)
        #         b = constants.extract_data( 'b', f)
        #     else:
        #         a = constants.extract_data('a', f )
        #         b = constants.extract_data('b', f)

        #     if 'GFDL-AM4' in directory or 'IPSL' in directory or 'GFDL-CM4' in directory:
        #         pass
        #     else:
        #         p0 = np.array(f.variables['p0'][:])
        #         a = a*p0

        # with Dataset( 'clw' + filename, 'r') as f:
        #     clw = np.nanmean( constants.extract_data( 'clw', f ), axis = 0 ) # average over time

        # with Dataset( 'cli' + filename, 'r') as f:
        #     cli = np.nanmean( constants.extract_data( 'cli', f ), axis = 0 ) # average over time

        # with Dataset( 'ps' + filename, 'r') as f:
        #     ps_lat_lon = np.mean( constants.extract_data( 'ps', f ), axis = 0 ) # average over time

        # with Dataset( 'ta' + filename, 'r' ) as f:
        #     ta = constants.extract_data( 'ta', f )
        #     ta[ta > 500] = np.nan
        #     ta = np.nanmean(ta, axis = 0) # average over time  
        #     plev_t = constants.extract_data('plev', f) / 100
        # full_ta_alt_lat = np.nanmean(ta, axis = -1)




    else:
        with Dataset( 'clt' + filename, 'r') as f:
            raw_lat = constants.extract_data( 'lat', f )
            raw_lon = constants.extract_data( 'lon', f )
            # clt = np.mean( constants.extract_data_over_time('clt', f, start, end ), axis = 0 ) # average over time

        # get indexes for southern ocean -70 to -50 lat

        start_idx = np.abs(raw_lat - (-69.5)).argmin()
        end_idx = np.abs(raw_lat - (-50)).argmin()


        #---get radiative flux data---#

        with Dataset( 'rsdt' + filename, 'r') as f:
            rsdt = np.mean( constants.extract_data_over_time('rsdt', f, start, end ), axis = 0 ) # average over time
            
        with Dataset( 'rsut' + filename, 'r') as f:
            rsut = np.mean( constants.extract_data_over_time('rsut', f, start, end ), axis = 0 ) # average over time           

        with Dataset( 'rsutcs' + filename, 'r') as f:
            rsutcs = np.mean( constants.extract_data_over_time('rsutcs', f, start, end ), axis = 0 ) # average over time           
        
        with Dataset( 'rlut' + filename, 'r') as f:
            rlut = np.mean( constants.extract_data_over_time('rlut', f, start, end ), axis = 0 ) # average over time           

        with Dataset( 'rlutcs' + filename, 'r') as f:
            rlutcs = np.mean( constants.extract_data_over_time('rlutcs', f, start, end ), axis = 0 ) # average over time           
        
        if 'MRI' in directory:
            rtmt = ( rsdt ) - ( rsut + rlut ) # https://www4.uwsp.edu/geo/faculty/lemke/geog101/lectures/02_radiation_energy_balance.html
        else:
            with Dataset( 'rtmt' + filename, 'r') as f:
                rtmt = np.mean( constants.extract_data_over_time('rtmt', f, start, end ), axis = 0 ) # average over time
        
                
        #---create albedo data---#
        
        albedo_reg = rsut / rsdt
        cre_sw_reg = rsutcs - rsut
        cre_lw_reg = rlutcs - rlut
        
        # with Dataset( 'cl' + filename, 'r') as f:
        #     cl = constants.extract_data_over_time( 'cl',f, start, end )
        #     cl = np.mean( cl, axis = 0 ) # average over time
        #     if directory == 'CMIP5-AMIP-CESM1-CAM5':
        #         pass
        #     else:
        #         cl = cl / 100
        #     if 'GFDL-AM4' in directory or 'IPSL' in directory or 'GFDL-CM4' in directory:
        #         a = constants.extract_data('ap', f)
        #         b = constants.extract_data( 'b', f)
        #     else:
        #         a = constants.extract_data('a', f )
        #         b = constants.extract_data('b', f)

        #     if 'GFDL-AM4' in directory or 'IPSL' in directory or 'GFDL-CM4' in directory:
        #         pass
        #     else:
        #         p0 = np.array(f.variables['p0'][:])
        #         a = a*p0

        # with Dataset( 'clw' + filename, 'r') as f:
        #     clw = np.nanmean( constants.extract_data_over_time( 'clw', f, start, end ), axis = 0 ) # average over time

        # with Dataset( 'cli' + filename, 'r') as f:
        #     cli = np.nanmean( constants.extract_data_over_time( 'cli', f, start, end ), axis = 0 ) # average over time

        # with Dataset( 'ps' + filename, 'r') as f:
        #     ps_lat_lon = np.mean( constants.extract_data_over_time( 'ps', f, start, end ), axis = 0 ) # average over time

        # with Dataset( 'ta' + filename, 'r' ) as f:
        #     ta = constants.extract_data_over_time( 'ta', f, start, end )
        #     ta[ta > 500] = np.nan
        #     ta = np.nanmean(ta, axis = 0) # average over time  
        #     plev_t = constants.extract_data('plev', f) / 100
        # full_ta_alt_lat = np.nanmean(ta, axis = -1)







    # a = np.swapaxes(np.swapaxes(np.tile(a, (np.shape(ps_lat_lon)[0], np.shape(ps_lat_lon)[1], 1) ), 0, 2), 1, 2)
    # b = np.swapaxes(np.swapaxes(np.tile(b, (np.shape(ps_lat_lon)[0], np.shape(ps_lat_lon)[1], 1) ), 0, 2), 1, 2)
    # ps = np.tile(ps_lat_lon, (np.shape(b)[0], 1, 1) )
    # p_3D = (a + b*ps) / 100 # in hPa
    
    # if directory == 'CMIP6-CESM2-CAM6':
    #     p_3D = np.flip(p_3D, axis = 0)

    # p_alt_lat = np.mean ( p_3D, axis = -1 )
    # p = constants.global3DMean(p_3D, raw_lat)
 
    # if directory == 'CMIP6-AMIP-CESM2-CAM6':
    #     cl_alt_lat = np.flip( np.nanmean( cl , axis = -1 ), axis = 0 ) # average over longitude
    #     clw_alt_lat = np.flip( np.nanmean( clw , axis = -1 ), axis = 0 ) # average over longitude
    #     cli_alt_lat = np.flip( np.nanmean( cli , axis = -1 ), axis = 0 ) # average over longitude    

    # else:
    #     cl_alt_lat = np.nanmean(cl , axis = -1) # average over longitude
    #     clw_alt_lat = np.nanmean(clw , axis = -1) # average over longitude
    #     cli_alt_lat = np.nanmean(cli , axis = -1) # average over longitude      


    #    #---Approximate missing nan values in temperature data---#  
       
    # imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    # imp.fit(np.transpose(full_ta_alt_lat))  
    # a = imp.transform(np.transpose(full_ta_alt_lat))
    # ta_alt_lat = np.transpose(a)    

    # if 'GFDL-CM4' in directory :
    #     interpolated = interpolate.interp2d(plev_t[:-1], raw_lat, np.transpose( ta_alt_lat ), kind = 'cubic')
    #     ta_alt_lat = np.transpose( interpolated(p, raw_lat) )
    # else:
    #     interpolated = interpolate.interp2d(plev_t, raw_lat, np.transpose( ta_alt_lat ), kind = 'cubic')
    #     ta_alt_lat = np.transpose( interpolated(p, raw_lat) )

    # #---create fractions---#

    # clw_frac = ( clw_alt_lat / ( clw_alt_lat + cli_alt_lat ) ) * cl_alt_lat

    # # calculate air density at each pressure layer
    # air_density_alt_lat = ((p_alt_lat * 100) / (286.9 * ta_alt_lat))

    # if 'IPSL-CM6A-LR' in directory:
    #     air_density_alt_lat = air_density_alt_lat[:-1]

    # clwc_alt_lat = ( clw_alt_lat * air_density_alt_lat ) * 1000 # in g/m3

    # store model name
    sheet.cell(row=rownum, column=1).value = directory
    # store clt
    # sheet.cell(row=rownum, column=2).value = constants.global2DMean(clt, raw_lat)
    # # store clt SO
    # sheet.cell(row=rownum, column=3).value = constants.global2DMean(clt[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # # store clwc
    # sheet.cell(row=rownum, column=4).value = constants.globalalt_latMeanVal(clwc_alt_lat, raw_lat)
    # # store clwc SO
    # sheet.cell(row=rownum, column=5).value = constants.globalalt_latMeanVal(clwc_alt_lat[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store SO albedo
    sheet.cell(row=rownum, column=6).value = constants.global2DMean(albedo_reg[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store net SO flux
    sheet.cell(row=rownum, column=7).value = constants.global2DMean(rtmt[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store SO CRE SW
    sheet.cell(row=rownum, column=8).value = constants.global2DMean(cre_sw_reg[start_idx:end_idx], raw_lat[start_idx:end_idx])
    # store SO CRE LW
    sheet.cell(row=rownum, column=9).value = constants.global2DMean(cre_lw_reg[start_idx:end_idx], raw_lat[start_idx:end_idx])


    book.save('E:/University/University/MSc/Models/climate-analysis/reduced_data/cloud_data.xlsx')

    # ######################---get profile data (time, plevel, lat, lon)---###################

    # with Dataset( 'cl' + filename, 'r') as f:
    #     cl = constants.extract_data_over_time( 'cl',f, start, end )
    #     cl = np.mean( cl, axis = 0 ) # average over time
    #     if directory == 'CMIP5-CESM1-CAM5':
    #        pass
    #     else:
    #         cl = cl / 100
            
    #     if directory == 'CMIP6-GFDL-AM4' or directory == 'CMIP5-IPSL-CM5A-LR' or directory == 'CMIP6-IPSL-CM6A-LR':
    #         a = constants.extract_data( f, 'ap')
    #         b = constants.extract_data( f, 'b')
    #     else:
    #         a = constants.extract_data( f, 'a')
    #         b = constants.extract_data( f, 'b')

    #     if directory == 'CMIP6-GFDL-AM4' or directory == 'CMIP5-IPSL-CM5A-LR' or directory == 'CMIP6-IPSL-CM6A-LR':
    #        pass
    #     else:
    #         p0 = np.array(f.variables['p0'][:])
    #         a = a*p0
    
    # with Dataset( 'clw' + filename, 'r') as f:
    #     clw = np.nanmean( constants.extract_data_over_time( 'clw', f, start, end ), axis = 0 ) # average over time

    # with Dataset( 'cli' + filename, 'r') as f:
    #     cli = np.nanmean( constants.extract_data_over_time( 'cli', f, start, end ), axis = 0 ) # average over time

    # with Dataset( 'ps' + filename, 'r') as f:
    #     ps_lat_lon = np.mean( constants.extract_data_over_time( 'ps', f, start, end ), axis = 0 ) # average over time

    # with Dataset( 'ta' + filename, 'r' ) as f:
    #     ta = constants.extract_data_over_time( 'ta', f, start, end )
    #     ta[ta > 500] = np.nan
    #     ta = np.nanmean(ta, axis = 0) # average over time  
    #     plev_t = constants.extract_data( f, 'plev') / 100



    # a = np.swapaxes(np.swapaxes(np.tile(a, (np.shape(ps_lat_lon)[0], np.shape(ps_lat_lon)[1], 1) ), 0, 2), 1, 2)
    # b = np.swapaxes(np.swapaxes(np.tile(b, (np.shape(ps_lat_lon)[0], np.shape(ps_lat_lon)[1], 1) ), 0, 2), 1, 2)
    # ps = np.tile(ps_lat_lon, (np.shape(b)[0], 1, 1) )
    # p_3D = (a + b*ps) / 100 # in hPa
    
    # if directory == 'CMIP6-CESM2-CAM6':
    #     p_3D = np.flip(p_3D, axis = 0)

    # p_alt_lat = np.mean ( p_3D, axis = -1 )

    # p_so = constants.global3DMean(p_3D[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
    # p = constants.global3DMean(p_3D, raw_lat)
    

    # #---convert pressure levels to altitude---#

    # #https://www.mide.com/pages/air-pressure-at-altitude-calculator
    # #https://www.grc.nasa.gov/www/k-12/airplane/atmosmet.html
    # raw_alt = np.empty((p.size,1),dtype=float)
    # state = 0
    # i = 0
    # for item in p:
    #     if state == 0:
    #         newalt = (288.19 - 288.08*((item/1012.90)**(1/5.256)))/6.49
    #         if newalt > 11:
    #             state = 1
    #     if state == 1:
    #         newalt = (1.73 - math.log(item/226.50))/0.157
    #         if( newalt > 25 ):
    #             state = 2
    #     if state == 2:
    #         newalt = (216.6*((item/24.88)**(1/-11.388)) - 141.94)/2.99
    #     raw_alt[i] = newalt
    #     i+=1

    # # get index for pressure just below 700hPa to define low cloud boundary
    # for p_index, a in enumerate (p):
    #     if a >= 700:
    #         index = p_index


    # raw_alt = np.transpose( raw_alt )[0]



    # #-------------create alt_lat grid sets---------------#

    # if directory == 'CMIP6-CESM2-CAM6':
    #     cl_alt_lat = np.flip( np.nanmean( cl , axis = -1 ), axis = 0 ) # average over longitude
    #     clw_alt_lat = np.flip( np.nanmean( clw , axis = -1 ), axis = 0 ) # average over longitude
    #     cli_alt_lat = np.flip( np.nanmean( cli , axis = -1 ), axis = 0 ) # average over longitude    
    #     cl = np.flip( cl, axis = 0 )  
    #     clw = np.flip( clw, axis = 0 )  

    # else:
    #     cl_alt_lat = np.nanmean(cl , axis = -1) # average over longitude
    #     clw_alt_lat = np.nanmean(clw , axis = -1) # average over longitude
    #     cli_alt_lat = np.nanmean(cli , axis = -1) # average over longitude      

    # full_ta_alt_lat = np.nanmean(ta, axis = -1)


    #    #---Approximate missing nan values in temperature data---#  
       
    # imp = SimpleImputer(missing_values=np.nan, strategy='mean')
    # imp.fit(np.transpose(full_ta_alt_lat))  
    # a = imp.transform(np.transpose(full_ta_alt_lat))
    # ta_fixed = np.transpose(a)    
    # #-------------create temp_alt grid sets---------------#


    # interpolated = interpolate.interp1d(p, raw_alt, fill_value="extrapolate", kind = 'cubic')
    # alt_temp = interpolated(plev_t)

    # if directory == 'CMIP6-IPSL-CM6A-LR':
    #     raw_alt = raw_alt[:79]
    #     interpolated = interpolate.interp1d(p[:40], raw_alt[:40], fill_value="extrapolate", kind = 'cubic')
    #     alt_temp = interpolated(plev_t)
    # else:
    #     interpolated = interpolate.interp1d(p, raw_alt, fill_value="extrapolate", kind = 'cubic')
    #     alt_temp = interpolated(plev_t)


    #   #-------------create cloud fraction datasets---------------#

    # # define low cloud cover (below 3km)
    # clt_lc_lat_lon = constants.lowregMean(cl[:index+1], p[:index+1])
    # clt_lc = constants.lowlatMean(cl[:index+1], p[:index+1])

    # cl_g = constants.global3DMean(cl, raw_lat)
    # cl_so = constants.global3DMean(cl[:,start_idx:end_idx], raw_lat[start_idx:end_idx])


    # #-------------create liquid and ice cloud fraction datasets---------------#

    # clwvi_lc = constants.lowlatMean(clw[:index+1], p[:index+1])
    # clwvi_lc_lat_lon = constants.lowregMean(clw[:index+1], p[:index+1])

    # clivi_lc = constants.lowlatMean(cli[:index+1], p[:index+1])
    # clivi_lc_lat_lon = constants.lowregMean(cli[:index+1], p[:index+1])

    # if directory == 'CMIP6-MIROC6':
    #     clwvi_lc_lat_lon = constants.fill(clwvi_lc_lat_lon)

    # clw_g = constants.global3DMean(clw, raw_lat)
    # clw_so = constants.global3DMean(clw[:,start_idx:end_idx], raw_lat[start_idx:end_idx])

    # cli_g = constants.global3DMean(cli, raw_lat)
    # cli_so = constants.global3DMean(cli[:,start_idx:end_idx], raw_lat[start_idx:end_idx])
        
    # #-------------interpolate liquid and ice fractions to common lat, alt and liq_alt---------------#


    # interpolated = interpolate.interp2d(raw_lat, raw_lon, np.transpose( clt_lc_lat_lon ), kind = 'cubic')
    # clt_lc_lat_lon = interpolated(constants.lat, constants.lon)

    # interpolated = interpolate.interp2d(raw_lat, raw_lon, np.transpose( clwvi_lc_lat_lon ), kind = 'cubic')
    # clwvi_lc_lat_lon = interpolated(constants.lat, constants.lon)

    # interpolated = interpolate.interp2d(raw_lat, raw_lon, np.transpose( clivi_lc_lat_lon ), kind = 'cubic')
    # clivi_lc_lat_lon = interpolated(constants.lat, constants.lon)

    # if directory == 'CMIP6-IPSL-CM6A-LR':
    #     cl_alt_lat = constants.fit_2d_data(np.transpose(cl_alt_lat), raw_lat, raw_alt)
    #     cl_alt_lat = constants.fill(cl_alt_lat)
    # else:
    #     interpolated = interpolate.interp2d(raw_alt, raw_lat, np.transpose( cl_alt_lat ), kind = 'cubic')
    #     cl_alt_lat = interpolated(constants.alt, constants.lat)

    # if directory == 'CMIP6-IPSL-CM6A-LR':
    #     interpolated = interpolate.interp2d(raw_alt[:40], raw_lat, np.transpose( clw_alt_lat[:40] ), kind = 'cubic')
    #     full_clw_alt_lat = interpolated(constants.alt, constants.lat)
    # else:
    #     interpolated = interpolate.interp2d(raw_alt, raw_lat, np.transpose( clw_alt_lat ), kind = 'cubic')
    #     full_clw_alt_lat = interpolated(constants.alt, constants.lat)

    # if directory == 'CMIP6-IPSL-CM6A-LR':
    #     interpolated = interpolate.interp2d(raw_alt[:40], raw_lat, np.transpose( cli_alt_lat[:40] ), kind = 'cubic')
    #     cli_alt_lat = interpolated(constants.alt, constants.lat)
    # else:    
    #     interpolated = interpolate.interp2d(raw_alt, raw_lat, np.transpose( cli_alt_lat ), kind = 'cubic')
    #     cli_alt_lat = interpolated(constants.alt, constants.lat)

    # interpolated = interpolate.interp1d(raw_lat, clt_lc, kind = 'cubic', fill_value="extrapolate")
    # clt_lc = interpolated(constants.lat)

    # interpolated = interpolate.interp1d(raw_lat, clwvi_lc, kind = 'cubic', fill_value="extrapolate")
    # clwvi_lc = interpolated(constants.lat)

    # interpolated = interpolate.interp1d(raw_lat, clivi_lc, kind = 'cubic', fill_value="extrapolate")
    # clivi_lc = interpolated(constants.lat)

    # interpolated = interpolate.interp1d(raw_alt, cl_g, kind = 'cubic', fill_value="extrapolate")
    # cl_g = interpolated(constants.alt)
    # if directory == 'CMIP6-IPSL-CM6A-LR' or directory == 'CMIP5-IPSL-CM5A-LR' or directory == 'CMIP6-MRI-ESM2':
    #     cl_g[:2] = np.nan
 

    # interpolated = interpolate.interp1d(raw_alt, clw_g, kind = 'cubic', fill_value="extrapolate")
    # clw_g = interpolated(constants.liq_alt)
    # clw_g[clw_g < 0] = np.nan

    # interpolated = interpolate.interp1d(raw_alt, cli_g, kind = 'cubic', fill_value="extrapolate")
    # cli_g = interpolated(constants.alt)
    # cli_g[cli_g < 0] = np.nan
    # cli_g[:2] = np.nan
    
    # interpolated = interpolate.interp1d(raw_alt, cl_so, kind = 'cubic', fill_value="extrapolate")
    # cl_so = interpolated(constants.alt)
    # if directory == 'CMIP6-IPSL-CM6A-LR' or directory == 'CMIP5-IPSL-CM5A-LR':
    #     cl_so[:2] = np.nan

    # interpolated = interpolate.interp1d(raw_alt, clw_so, kind = 'cubic', fill_value="extrapolate")
    # clw_so = interpolated(constants.liq_alt)
    # clw_so[clw_so < 0] = np.nan

    # interpolated = interpolate.interp1d(raw_alt, cli_so, kind = 'cubic', fill_value="extrapolate")
    # cli_so = interpolated(constants.alt)
    # cli_so[cli_so < 0] = np.nan
    # cli_so[:2] = np.nan


    # #-------------interpolate temp data to common lat, alt and liq_alt---------------#

    # interpolated = interpolate.interp2d(alt_temp, raw_lat, np.transpose( ta_fixed ), kind = 'cubic')
    # full_ta_alt_lat = interpolated(constants.alt, constants.lat)

    # if directory == 'CMIP6-IPSL-CM6A-LR':
    #     interpolated = interpolate.interp2d(raw_alt[:40], raw_lat, np.transpose( p_alt_lat[:40] ), kind = 'cubic')
    #     full_p_alt_lat = interpolated(constants.alt, constants.lat)
    # else:    
    #     interpolated = interpolate.interp2d(raw_alt, raw_lat, np.transpose( p_alt_lat ), kind = 'cubic')
    #     full_p_alt_lat = interpolated(constants.alt, constants.lat)

    # #-------------calculate density---------------#


    # ta_alt_lat = full_ta_alt_lat[:,:constants.liq_alt_confine]
    # clw_alt_lat = full_clw_alt_lat[:,:constants.liq_alt_confine]
    # p_alt_lat = full_p_alt_lat[:,:constants.liq_alt_confine]


    # # calculate air density at each pressure layer
    # full_air_density_alt_lat = ((full_p_alt_lat * 100) / (286.9 * full_ta_alt_lat))
    # liq_air_density_alt_lat = ((p_alt_lat * 100) / (286.9 * ta_alt_lat))

    # full_air_density_g = constants.globalalt_latMean( np.transpose( full_air_density_alt_lat ), constants.lat ) # corresponding to (alt)
    # full_air_density_so = constants.globalalt_latMean( np.transpose( full_air_density_alt_lat[constants.so_idx_1:constants.so_idx_2] ), constants.lat[constants.so_idx_1:constants.so_idx_2] ) # corresponding to (alt)

    # liq_air_density_g = constants.globalalt_latMean( np.transpose( liq_air_density_alt_lat ), constants.lat ) # corresponding to (liq_alt)
    # liq_air_density_so = constants.globalalt_latMean( np.transpose( liq_air_density_alt_lat[constants.so_idx_1:constants.so_idx_2] ), constants.lat[constants.so_idx_1:constants.so_idx_2] ) # corresponding to (liq_alt)

    # full_clwc_alt_lat = ( full_clw_alt_lat * full_air_density_alt_lat ) * 1000 # in g/m3
    # clwc_alt_lat = ( clw_alt_lat * liq_air_density_alt_lat ) * 1000 # in g/m3
    # clwc_alt_lat[ clwc_alt_lat < 0 ] = None

    # clwc_g = ( clw_g * liq_air_density_g ) * 1000 # in g/m3
    # clwc_so = ( clw_so * liq_air_density_so ) * 1000 # in g/m3

    # # fig, ax = plt.subplots()
    # # cont = ax.contourf( constants.lat, constants.liq_alt, np.transpose(p_alt_lat) )
    # # temp = ax.contour( constants.lat, constants.liq_alt, (np.transpose(ta_alt_lat) - 273.15), colors='white')
    # # temp.collections[5].set_linewidth(3)
    # # temp.collections[5].set_color('white')
    # # ax.clabel(temp, inline=1, fontsize=10)
    # # ax.set_xlabel('Latitude')
    # # ax.set_ylabel('Altitude (km)')
    # # cbar = fig.colorbar(cont, orientation='horizontal')
    # # plt.show()

    #  ######## Binned Temperature Data ########
    # # values to bin: clw_alt_lat and ta_alt_lat
    # # binned into constants.ta_g array
    # # values in each bin to be summed
    # # call the summed values clw_t_g

    # stat = 'mean'

    # cl_t_g, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat.flatten(), cl_alt_lat.flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))
    # cl_t_so, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), cl_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))

    # clw_t_g, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat.flatten(), full_clw_alt_lat.flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))
    # clw_t_so, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), full_clw_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))

    # clwc_t_g, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat.flatten(), full_clwc_alt_lat.flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))
    # clwc_t_so, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), full_clwc_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))

    # cli_t_g, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat.flatten(), cli_alt_lat.flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))
    # cli_t_so, bin_edges, binnumber = stats.binned_statistic(full_ta_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), cli_alt_lat[constants.so_idx_1:constants.so_idx_2].flatten(), stat, bins=constants.ta.size, range=(constants.min_ta, constants.max_ta))

    # clw_frac_t_g = ( clw_t_g / ( clw_t_g + cli_t_g ) ) * cl_t_g
    # clw_frac_t_so = ( cli_t_g / ( clw_t_g + cli_t_g ) ) * cl_t_g

    # # fig, ax = plt.subplots()
    # # ax.plot( constants.ta, clw_t_g )
    # # ax.plot( constants.ta, clw_t_so )
    # # ax.axvline(x=273, label = '273K', color = 'black', linestyle='--')
    # # plt.grid(True)
    # # plt.show()
    # ######################


    # #---create fractions---#

    # full_clw_frac_alt_lat = ( full_clw_alt_lat / ( full_clw_alt_lat + cli_alt_lat ) ) * cl_alt_lat
    # cli_frac_alt_lat = ( cli_alt_lat / ( full_clw_alt_lat + cli_alt_lat ) ) * cl_alt_lat
    # clw_frac_alt_lat = full_clw_frac_alt_lat[:,:constants.liq_alt_confine]


    # os.chdir(save_location)

    # save_filename = start.strftime( '%b_%Y') + '_' + end.strftime( '%b_%Y') + '_' + directory + '.h5'

    # with h5py.File(save_filename, 'w') as p:
        
    #     p.create_dataset('clt', data=clt) # total cloud fraction corresponding to lat
    #     p.create_dataset('clt_lc', data=clt_lc) # low cloud fraction corresponding to lat
    #     p.create_dataset('clt_lat_lon', data=np.transpose( clt_lat_lon )) # total cloud fraction corresponding to lat, lon
    #     p.create_dataset('clt_lc_lat_lon', data=np.transpose( clt_lc_lat_lon )) # low cloud fraction corresponding to lat, lon
 
    #     p.create_dataset('clwvi', data=clwvi) # total cloud liquid water fraction corresponding to lat
    #     p.create_dataset('clwvi_lc', data=clwvi_lc) # low cloud fraction corresponding to lat
    #     p.create_dataset('clwvi_lat_lon', data=np.transpose( clwvi_lat_lon )) # total cloud liquid water fraction corresponding to lat, lon
    #     p.create_dataset('clwvi_lc_lat_lon', data=np.transpose( clwvi_lc_lat_lon )) # low cloud liquid water fraction corresponding to lat, lon

    #     p.create_dataset('clivi', data=clivi) # total cloud ice fraction corresponding to lat
    #     p.create_dataset('clivi_lc', data=clivi_lc) # low cloud fraction corresponding to lat
    #     p.create_dataset('clivi_lat_lon', data=np.transpose( clivi_lat_lon )) # total cloud fraction corresponding to lat, lon
    #     p.create_dataset('clivi_lc_lat_lon', data=np.transpose( clivi_lc_lat_lon )) # low cloud liquid water fraction corresponding to lat, lon
      
    #     p.create_dataset('cl_g', data=cl_g) # global layer total cloud fraction corresponding to alt
    #     p.create_dataset('clw_g', data=clw_g) # global layer cloud liquid water fraction(kg/kg) corresponding to liq_alt
    #     p.create_dataset('cli_g', data=cli_g) # global layer cloud ice water fraction(kg/kg) corresponding to alt
    #     p.create_dataset('clwc_g', data=clwc_g) # global layer cloud liquid water content (g/m3) corresponding to liq_alt
    
    #     p.create_dataset('cl_so', data=cl_so) # southern ocean layer total cloud fraction corresponding to alt
    #     p.create_dataset('clw_so', data=clw_so) # southern ocean layer cloud liquid water fraction(kg/kg) corresponding to liq_alt
    #     p.create_dataset('cli_so', data=cli_so) # southern ocean layer cloud ice water fraction(kg/kg) corresponding to alt
    #     p.create_dataset('clwc_so', data=clwc_so) # southern ocean layer cloud liquid water content (g/m3) corresponding to liq_alt
 
    #     p.create_dataset('cl_t_g', data=cl_t_g) # global layer cloud fraction corresponding to ta
    #     p.create_dataset('cl_t_so', data=cl_t_so) # global layer cloud fraction corresponding to ta

    #     p.create_dataset('clw_t_g', data=clw_t_g) # global layer cloud liquid water fraction (kg/kg) corresponding to ta
    #     p.create_dataset('clw_t_so', data=clw_t_so) # global layer cloud liquid water fraction (kg/kg) corresponding to ta

    #     p.create_dataset('clw_frac_t_g', data=clw_frac_t_g) # global layer cloud liquid water fraction corresponding to ta
    #     p.create_dataset('clw_frac_t_so', data=clw_frac_t_so) # global layer cloud liquid water fraction corresponding to ta

    #     p.create_dataset('clwc_t_g', data=clwc_t_g) # global layer cloud liquid water content (g/m3) corresponding to ta
    #     p.create_dataset('clwc_t_so', data=clwc_t_so) # global layer cloud liquid water content (g/m3) corresponding to ta

    #     p.create_dataset('ta_alt_lat', data=np.transpose( ta_alt_lat )) # temperature (K) corresponding to liq_alt and lat
    #     p.create_dataset('cl_alt_lat', data=np.transpose( cl_alt_lat ) ) # total cloud fraction corresponding to alt and lat
    #     p.create_dataset('clw_alt_lat', data=np.transpose( clw_alt_lat ) ) # cloud liquid water fraction (kg/kg) corresponding to liq_alt and lat
    #     p.create_dataset('cli_alt_lat', data=np.transpose( cli_alt_lat ) ) # cloud ice water fraction (kg/kg) corresponding to alt and lat
    #     p.create_dataset('clwc_alt_lat', data=np.transpose( clwc_alt_lat ) ) # cloud liquid water content (g/m3) corresponding to liq_alt and lat

    #     p.create_dataset('full_clw_frac_alt_lat', data=np.transpose( full_clw_frac_alt_lat ) ) # cloud liquid water fraction (kg/kg) corresponding to liq_alt and lat
    #     p.create_dataset('clw_frac_alt_lat', data=np.transpose( clw_frac_alt_lat ) ) # cloud liquid water fraction (kg/kg) corresponding to liq_alt and lat
    #     p.create_dataset('cli_frac_alt_lat', data=np.transpose( cli_frac_alt_lat ) ) # cloud ice water fraction (kg/kg) corresponding to alt and lat

    #     p.create_dataset('full_ta_alt_lat', data= np.transpose( full_ta_alt_lat )) # temperature corresponding to alt and lat
    #     p.create_dataset('full_clw_alt_lat', data=np.transpose( full_clw_alt_lat ) ) # cloud liquid water fraction (kg/kg) corresponding to alt and lat
        
    #     p.create_dataset('rtmt_lat_lon', data=np.transpose( rtmt_lat_lon ) ) # net downwards raditaive flux (W/m2) at toa corresponding to lat, lon
    #     p.create_dataset('albedo_reg', data=np.transpose( albedo_reg ) ) 
    #     p.create_dataset('cre_sw_reg', data=np.transpose( cre_sw_reg ) ) # cloud radiative effect (clear sky - all sky) shortwave
    #     p.create_dataset('cre_lw_reg', data=np.transpose( cre_lw_reg ) ) # cloud radiative effect (clear sky - all sky) longwave
 
    #     p.create_dataset('full_air_density_alt_lat', data=np.transpose( full_air_density_alt_lat ) ) # kg/m3
    #     p.create_dataset('liq_air_density_alt_lat', data=np.transpose( liq_air_density_alt_lat ) ) # kg/m3

    #     if directory == 'CMIP6-GFDL-AM4':
    #         p.create_dataset('mmrdust_lat_lon', data=np.transpose( mmrdust_lat_lon )) # aerosol dust mixing ratio corresponding to lat, lon
    #         p.create_dataset('mmroa_lat_lon', data=np.transpose( mmroa_lat_lon ) ) # aerosol total organic mixing ratio corresponding to lat, lon
    #         p.create_dataset('mmrso4_lat_lon', data=np.transpose( mmrso4_lat_lon ) ) # aerosol so4 mixing ratio corresponding to lat, lon
    #         p.create_dataset('aerosol_norm_lat_lon', data=np.transpose( aerosol_norm_lat_lon ) ) # total normalised aerosols corresponding to lat, lon

    #     if directory == 'CMIP5-GISS-E2R' or directory == 'CMIP5-MIROC5':
    #         p.create_dataset('mmrdust_lat_lon', data=np.transpose( mmrdust_lat_lon )) # aerosol dust mixing ratio corresponding to lat, lon
    #         p.create_dataset('mmroa_lat_lon', data=np.transpose( mmroa_lat_lon ) ) # aerosol total organic mixing ratio corresponding to lat, lon
    #         p.create_dataset('mmrso4_lat_lon', data=np.transpose( mmrso4_lat_lon ) ) # aerosol so4 mixing ratio corresponding to lat, lon
    #         p.create_dataset('mmrbc_lat_lon', data=np.transpose( mmrbc_lat_lon ) ) # aerosol total organic mixing ratio corresponding to lat, lon
    #         p.create_dataset('mmrss_lat_lon', data=np.transpose( mmrss_lat_lon ) ) # aerosol so4 mixing ratio corresponding to lat, lon
    #         p.create_dataset('aerosol_norm_lat_lon', data=np.transpose( aerosol_norm_lat_lon ) ) # total normalised aerosols corresponding to lat, lon


    #     p.close()
